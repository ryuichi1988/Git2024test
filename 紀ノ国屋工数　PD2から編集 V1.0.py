"""
2021.10.03
食事修正
コード３４６行目
食事１８１→１８５行に修正。
ファイル：小田原NEWマスター

2021.12.06
v4.0
個人食数集計できるようになった。
退職者を消した。
食事185行→145行に修正。

2022.1.10
V4.1
2022年対応済み

2022.5.16
V4.4

下記修正済み：（どっちの○を書いても、強制的にただしい○にする。
143        if syokuji == "○" or syokuji == "〇":
144            syokuji = "○"

2022.6.12
v4.4.2
食事145行→126行に修正。

2022.7.4
v4.4.3
退職者を削除したため、行数減少。
食事126行→108行に修正。

2022.12.12
v4.4.3
食事146に修正。


2022.12.25
夜勤20時対応済み。
遅刻者については、正しく処理されません。→20:15など
もう作業する必要ないので、修正は見送ります。

2023.1.2
あけおめ！ことよろ。ｗ
ロードする時、ファイル名を2023年に変更しました。

2024.1.6 update
明けましておめでとうございます。
今年もよろしくお願いします！！
いやー、時間過ぎるのは早いっすね(笑)

2024対応するようにします。
4-5箇所かな。
検索機能使いました。


2024.3.5 UPDATE
食事⇒158


2024.7.16
プライム第二仕様にしました。

2024.12.02
東京紀ノ国屋様仕様に変更。
マニュアル：
pathMEIBO = r"C:\1\PD2\紀ノ国屋様出勤者名簿.xlsx"
pathKOSU = r"C:\1\PD2\紀ノ国屋様工数.xlsx"


"""

import openpyxl
import difflib
from sys import exit
from datetime import timedelta, datetime, time
from tkinter import Tk,messagebox

root = Tk()
root.attributes('-topmost', True)
root.withdraw()

pathMEIBO = r"C:\1\PD2\紀ノ国屋様出勤者名簿.xlsx"
pathKOSU = r"C:\1\PD2\紀ノ国屋様工数.xlsx"

def format_timedelta(timedelta):
  total_sec = timedelta.total_seconds()
  # hours
  hours = total_sec // 3600 
  # remaining seconds
  remain = total_sec - (hours * 3600)
  # minutes
  minutes = remain // 60
  # remaining seconds
#   seconds = remain - (minutes * 60)
  # total time
  return '{:02}:{:02}'.format(int(hours), int(minutes))

while True:
    allWorkTimeToday = timedelta(hours=0)
    n = 0

    # 食数カウント用　I列の"○"をカウントする。
    syokujicount = 0



    SAGYOUDETA = input("作業データの日付を入力してください。（４桁数字、例：1021）終了するにはENTERキーを押してください。")
    # もしエンターキーが押されたら終了。
    if SAGYOUDETA == "":
        exit()

    dcmonth = SAGYOUDETA[0:2]  # input月
    dcday = SAGYOUDETA[2:]  # input日

    # 工数集計表を開く
    try:
        PrimeDai2MeiBoFile = openpyxl.load_workbook(pathMEIBO)
        PrimeDai2todaySheet = PrimeDai2MeiBoFile["{}".format(SAGYOUDETA)]
        print("既存データ{}.xlsxを開きました,少々お待ち下さい。".format(SAGYOUDETA))
    except KeyError :    
        ret = messagebox.showinfo('ERROR message', '{}のデータが存在しません。'.format(SAGYOUDETA))   #
        exit()
            

    # NEWマスターファイルを開く (10月分)
    PD2kosuEXCEL = openpyxl.load_workbook(pathKOSU)
    PD2kosuExcel_Book0SHEET = PD2kosuEXCEL["0SHEET"]

        # difflib ready

    diffsheet = PD2kosuEXCEL["2在籍者名簿Master"]

    # NEW*月分勤怠表を開く(10月分)
    # PD2kosuEXCEL = openpyxl.load_workbook(r"C:\1\py\NEW{}月分勤怠表奥山ver.xlsx".format(dcmonth))

    # 入力待ち

    # エクセル時間を取り込み
    # try:
    #     kousuudeta = openpyxl.load_workbook(r"C:\1\py\実績\【2024{}実績】ダイエットクック日別勤怠集計表奥山ver2.xlsx".format(SAGYOUDETA))
    # except FileNotFoundError:
    #     print("32実績データエラー、お確かめの上、最初からやり直してください。")
    #     exit()

    for row in range(5, 50, 1):
        compareinilist = []
        # B列の名前はnamenow
        namenow = PrimeDai2todaySheet.cell(row=row, column=3).value
        if namenow is None or namenow == "" or namenow == " " or namenow =="　":
            print(namenow)
            print("continue 144")
            continue
        elif str(namenow).isdigit():
            print("B列エラー、確認してください。プログラム終了")
            exit()
        print("入力された名前" + str(namenow))
        # diff操作
        for row2 in range(1, len(diffsheet["A"]) + 1):

            nameForCompared = diffsheet.cell(row=row2, column=2).value  ##ここでエラーだと、空白行があること。
            try:
                compareini = difflib.SequenceMatcher(None, namenow, nameForCompared).quick_ratio()
            except AttributeError:
                continue
            compareinilist.append(compareini)  # 比較したマッチ度をリストに追加
        highestini = max(compareinilist)  # MAXマッチ度
        highestiniNo = compareinilist.index(highestini)  # INDEX
        highestiniName = diffsheet.cell(row=highestiniNo + 1, column=2).value  # 名前
        staffnumber = diffsheet.cell(row=highestiniNo + 1, column=1).value  # NP番号です。
        # print("マッチした名前" + str(highestiniName) + "  番号:" + str(staffnumber))

        # 実績データに一応記入する。１名前、２番号　

        highestini = highestini * 100
        highestini = str(round(highestini, 1)) + "%"

        PrimeDai2todaySheet.cell(row=row, column=4).value = staffnumber
        PrimeDai2todaySheet.cell(row=row, column=2).value = highestiniName
        PrimeDai2todaySheet.cell(row=row, column=1).value = highestini

        # 実績データに一応記入する。１名前、２番号　

        kousuunumbernow = staffnumber  # 実績データK列ｐｙ番号
        if kousuunumbernow is None:
            print("シート取り込み完了")
            break
        else:
            pass
        kousuustarttimeraw = PrimeDai2todaySheet.cell(row=row, column=8).value  # F列始業時間
        kousuuendtimeraw = PrimeDai2todaySheet.cell(row=row, column=9).value
        # 時間処理
        if type(kousuustarttimeraw) == timedelta :
            pass
        elif type(kousuustarttimeraw ) == time :
            kousuustarttimeraw = timedelta(hours=kousuustarttimeraw.hour,minutes=kousuustarttimeraw.minute)
            if type(kousuuendtimeraw) == timedelta:
                pass
            else:
                kousuuendtimeraw = timedelta(hours=kousuuendtimeraw.hour,minutes=kousuuendtimeraw.minute)
        else:
            # messagebox.showinfo('170エラー', '170エラー')   #「情報」のメッセージボックスを表示
            print(type(kousuustarttimeraw))
            continue

        # try:

        #     print(kousuustarttimeraw)
        #     print(type(kousuustarttimeraw))
        #     starttime = timedelta(hours=kousuustarttimeraw.hour, minutes=kousuustarttimeraw.minute)
        # except AttributeError:
        #     print("エラー発生。5行目は時間ではありませんでした。プログラム終了")
            
          # G列就業時間

        # 　時間処理
        print(kousuuendtimeraw)
        print(type(kousuuendtimeraw))
        starttime = kousuustarttimeraw
        endtime = kousuuendtimeraw
        # endtime = timedelta(hours=kousuuendtimeraw.hour, minutes=kousuuendtimeraw.minute)
        if starttime > endtime:
        # if starttime == timedelta(hours=19) or starttime == timedelta(hours=20) or starttime == timedelta(hours=20,minutes=30):
            endtime = endtime + timedelta(hours=24)

        kousuuresttimeraw = PrimeDai2todaySheet.cell(row=row, column=10).value  # H列休憩時間
        # 休憩時間処理
        if type(kousuuresttimeraw) == timedelta:
            pass
            resttime = kousuuresttimeraw
        else:
            print("休憩なし")
            resttime = timedelta(hours=0)
        # 食数表示
        syokuji = PrimeDai2todaySheet.cell(row=row, column=11).value
        print("食事表示： {0}".format(syokuji))
        if syokuji == "○" or syokuji == "〇":
            syokuji = "○"
            syokujicount += 1
            print("食事有り：○")
        else:
            print("食事無し")
        # データ取り込み完了

        PrimeDai2todaySheet["K1"] = "{}月{}日".format(dcmonth, dcday)
        # print("{}月{}日工数入力開始".format(dcmonth, dcday))
        # userinput = str(kousuunumbernow)+str()
        #    userinput = input("6-13桁指令。ENTERで終了")
        # no = input("番号を入力してください（４桁数字）。空白値で終了")
        #   if userinput == "":
        #       break
        #     elif len(userinput) != 6 and len(userinput) != 13:
        #         print("入力に間違いがあります。お確かめの上、もう一度入力してください。")
        #        continue
        #    else:
        #       pass
        # 番号を7000番台に変換
        #   kousuunumbernow = int(userinput[0:4]) + 70000000

        # if len(userinput) == 6 and userinput[-2:] == "08":
        #    starttimeH = int(8)
        #    starttimeM = int(30)
        #    starttime = timedelta(hours=starttimeH, minutes=starttimeM)

        # 就業時間処理
        #    endtimeH = int(17)
        #    endtimeM = int(30)
        #    endtime = timedelta(hours=endtimeH, minutes=endtimeM)

        # 休憩時間処理
        #    resttimeH = int(1)
        #    resttime = timedelta(hours=resttimeH)

        # elif len(userinput) == 6 and userinput[-2:] == "10":
        #     starttimeH = int(10)
        #     starttimeM = int(00)
        #     starttime = timedelta(hours=starttimeH, minutes=starttimeM)
        #
        #     # 就業時間処理
        #     endtimeH = int(19)
        #     endtimeM = int(00)
        #     endtime = timedelta(hours=endtimeH, minutes=endtimeM)
        #
        #     # 休憩時間処理
        #     resttimeH = int(1)
        #     resttime = timedelta(hours=resttimeH)
        #
        #
        # elif len(userinput) == 6 and userinput[-2:] == "18":
        #     starttimeH = int(18)
        #     starttimeM = int(00)
        #     starttime = timedelta(hours=starttimeH, minutes=starttimeM)
        #
        #     # 就業時間処理
        #     endtimeH = int(24)
        #     endtimeM = int(00)
        #     endtime = timedelta(hours=endtimeH, minutes=endtimeM)
        #
        #     # 休憩時間処理
        #     resttimeH = int(0)
        #     resttime = timedelta(hours=resttimeH)
        #
        # else:
        #     # 始業時間処理
        #     starttimeH = int(userinput[4:6])
        #     starttimeM = int(userinput[6:8])
        #     starttime = timedelta(hours=starttimeH, minutes=starttimeM)
        #
        #     # 就業時間処理
        #     endtimeH = int(userinput[8:10])
        #     endtimeM = int(userinput[10:12])
        #     endtime = timedelta(hours=endtimeH, minutes=endtimeM)
        #
        #     # 休憩時間処理
        #     resttimeH = int(userinput[-1])
        #     resttime = timedelta(hours=resttimeH)

        # 休憩開始、終了時間処理
        # # もし出勤時間が5時と10時の間且つ休憩が1時間の場合：
        # if starttime >= timedelta(hours=5) and starttime < timedelta(hours=10) and resttime == timedelta(hours=1):
        #     resttimestart = timedelta(hours=12)
        #     resttimefinish = timedelta(hours=13)
        # # もし出勤時間が10時00分且つ休憩が1時間の場合
        # elif starttime >= timedelta(hours=10) and starttime < timedelta(hours=12) and resttime == timedelta(hours=1):
        #     resttimestart = timedelta(hours=13)
        #     resttimefinish = timedelta(hours=14)
        # # 上記2パターン以外]
        # elif starttime == timedelta(hours=19) and resttime == timedelta(hours=1):
        #     resttimestart = timedelta(hours=22,minutes=45)
        #     resttimefinish = timedelta(hours=23,minutes=45)

        # elif starttime == timedelta(hours=20) and resttime == timedelta(hours=1):
        #     resttimestart = timedelta(hours=22,minutes=45)
        #     resttimefinish = timedelta(hours=23,minutes=45)
        # 休憩時間処理
        if resttime == timedelta(hours=1):

            if starttime < timedelta(hours=10):
                resttimestart = timedelta(hours=12)
                resttimefinish = timedelta(hours=13)
            else:
                resttimestart = timedelta(hours=0)
                resttimefinish = timedelta(hours=1)
        else:
            # starttime >= timedelta(hours=18) or starttime == timedelta(hours=19) or resttime == timedelta(hours=0):
            resttimestart = timedelta(hours=0)
            resttimefinish = timedelta(hours=0)
        # else:
        #     print("休憩時間エラー")
        #     continue

        # 労働時間処理
        worktime = endtime - starttime - resttime
        if worktime < timedelta(hours=0):
            worktime = worktime + timedelta(hours=24)
        else:
            pass


        PrimeDai2todaySheet.cell(row=row, column=5).value = worktime
        PrimeDai2todaySheet.cell(row=row, column=5).number_format = "[h]:mm"

        allWorkTimeToday = allWorkTimeToday + worktime

        # print("{}\n{}\n{}\n{}".format(starttime,endtime,resttime,worktime))



        # この操作は不要と思います。↓


        # 番号特定作業、時間入力
        # B列の番号を探す
        # for cell in PrimeDai2todaySheet["B"]:
        #     # もし入力された番号がヒットしたら
        #     if cell.value == kousuunumbernow:
        #         # まずは名前を表示する
        #         name = PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 1).value
        #         # 確認用
        #         print("名前【 {} 】\n実働【 {} 】\n休憩【 {} 】時間".format(name, str(worktime)[:-3], str(resttime)[0]))
        #         # userinput = input("データを入力してください。（9桁数字）例：080019001　←時刻数字開始+終了+最後の数字は休憩時間")
        #         # starttime = userinput[:4]
        #         # endtime = userinput[4:9]
        #         # resttime = userinput[-1]
        #         # worktime = 0
        #         # 本番、時間入力
        #         # try:
        #         PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 2).value = starttime
        #         # except PermissionError:
        #         #     print("エラー。エクセルを閉じてからこのプログラムを実行してください。")
        #         #     break
        #         if endtime == timedelta(hours=24, minutes=00) or endtime == timedelta(hours=00, minutes=00):
        #             # もし０時退勤：
        #             PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 5).value = "24:00"
        #             PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 5).number_format = "[h]:mm"
        #         else:
        #             # ０時退勤ではない
        #             PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 5).value = endtime
        #         # 休憩時間処理
        #         if resttime == timedelta(hours=00, minutes=00):
        #             # もし休憩無し：
        #             PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 8).value = str('')
        #         else:
        #             # 休憩がゼロ以外；
        #             PrimeDai2todaySheet.cell(row=cell.row, column=cell.column + 8).value = resttime

        #         break

        # 小田原NEWマスタにデータを入れる。
        for cell in PD2kosuExcel_Book0SHEET["C"]:
            # もし入力された番号がヒットしたら
            if cell.value == kousuunumbernow:
                # まずは名前を表示する
                # name = PD2kosuExcel_Book0SHEET.cell(row=cell.row, column=cell.column + 1).value
                # print("NEWマスタ名前特定：{},{}".format(name, cell))
                todaycell = PD2kosuExcel_Book0SHEET.cell(row=cell.row, column=int(dcday) + 4)
                todaycell.value = worktime
                todaylunchsell = PD2kosuExcel_Book0SHEET.cell(row=cell.row, column=int(dcday) + 44)  # 作業日付からカウント、+40＝食事セル
                if syokuji == "○":
                    todaylunchsell.value = 1  # 当日の食事を１にする。
                else:
                    break

            # print("NEWマスタ名前{}特定できませんでした。確認してください。".format(kousuunumbernow))

        # NEW*月分勤怠表（個人詳細）にデータ入れる
        # シート名前遍歴
        list = PD2kosuEXCEL.sheetnames
        for i in list:
            # もし番号がシート名に含まれている場合
            if str(kousuunumbernow) in i:
                # シート名
                # print(i)
                PD2kosuExcel_PersonalSheet = PD2kosuEXCEL[i]
                # 特定したシートのD4値（確認用）
                # print(PD2kosuExcel_PersonalSheet["D4"].value)
                # 出勤時間を入れる
                PD2kosuExcel_PersonalSheet["C{}".format(int(dcday) + 8)].value = starttime
                # 退勤時間を入れる
                if endtime == timedelta(hours=24, minutes=00):
                    PD2kosuExcel_PersonalSheet["E{}".format(int(dcday) + 8)].value = "24:00"
                    PD2kosuExcel_PersonalSheet["E{}".format(int(dcday) + 8)].number_format = "[h]:mm"
                else:
                    PD2kosuExcel_PersonalSheet["E{}".format(int(dcday) + 8)].value = endtime
                # 休憩開始時間を入れる
                if resttime == timedelta(hours=00, minutes=00):
                    pass
                else:
                    PD2kosuExcel_PersonalSheet["F{}".format(int(dcday) + 8)].value = resttimestart
                    # 休憩終了時間を入れる
                    PD2kosuExcel_PersonalSheet["G{}".format(int(dcday) + 8)].value = resttimefinish
                # 　食事数を入れる。
                if syokuji == "○":
                    PD2kosuExcel_PersonalSheet["O{}".format(int(dcday) + 8)].value = 1
                break
            # print("{}のシートが見つかりません。確認してください。".format(i))
        # PD2kosuExcel_PersonalSheet = PD2kosuEXCEL.active

        # print("保存済\n")

        # NEW*月分勤怠表を入れる
        n += 1
        print("{}人入力完成 \n----------------------------------".format(n))

    allWorkTimeTodayHH = "{}時間 ".format((allWorkTimeToday.days * 24 + allWorkTimeToday.seconds / 3600))
    allWorkTimeTodayHH2 = format_timedelta(allWorkTimeToday)
    word = "日付：{0}　人数：{1}人,　食数：{2}食, 100進時間：{3}　（60進時間：{4}) ".format(SAGYOUDETA, n, syokujicount, allWorkTimeTodayHH,allWorkTimeTodayHH2)
    print(word)
    PrimeDai2todaySheet["A2"] = word

    # # 幅を指定
    PrimeDai2todaySheet.column_dimensions["B"].width = 20
    # PrimeDai2todaySheet.column_dimensions["A"].width = 8
    PrimeDai2todaySheet.column_dimensions["D"].width = 11
    PrimeDai2todaySheet.column_dimensions["E"].width = 11

    # # 非表示にする：
    # for row in range(2, 9, 1):
    #     PrimeDai2todaySheet.row_dimensions[row].hidden = True

    # PrimeDai2todaySheet.column_dimensions["J"].hidden = True
    # PrimeDai2todaySheet.column_dimensions["K"].hidden = True

    # 幅を指定完了
    PrimeDai2todaySheet["A3"] = "マッチ度"
    PrimeDai2todaySheet["B4"] = "ﾊﾟｲｿﾝが認識した名前"
    PrimeDai2todaySheet["D3"] = "NP番号"
    PrimeDai2todaySheet["E3"] = "実働時間"

    print("ファイルを保存しています・・・")
    # 食数を入力する：　小田原NEW　EXCEL　１８５行目、列はDCDAY　＋　４　（例３日の場合：７列目（G））
    PD2kosuExcel_Book0SHEET.cell(row=158, column=int(dcday) + 4).value = syokujicount
    PrimeDai2MeiBoFile.save(pathMEIBO.format(SAGYOUDETA))
    PD2kosuEXCEL.save(pathKOSU.format(dcmonth))