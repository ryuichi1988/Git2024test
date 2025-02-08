import re
import fitz
import openpyxl
from paddleocr import PaddleOCR
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from tqdm import tqdm


# ============================ 初始化数据 ============================#
ocr = PaddleOCR(
    use_angle_cls=False,
    lang='japan',
)
MAX_X_DISTANCE = 350  # 允许的最大 X 坐标距离，超过则认为不是同一行
THRESHOLD_Y = 5  # 允许的 Y 坐标误差
current_line_times = []
nowtime = datetime.now()
wb = openpyxl.load_workbook("PDtestM.xlsm", keep_vba=True)
source = wb["99999　ニッセープロダクツ"]
number_master_sheet = wb['Number_Master']
number_name_dict = {}
name_now = None
current_date = None
last_name = None  # 记录名字
data_list = []  # 用于收集所有数据，直到姓名发生变化时才写入 Excel
# ============================ 初始化数据 ============================#



# ============================ 写入Master字典 ============================#
for row in range(1, number_master_sheet.max_row + 1):
    PD_number = number_master_sheet.cell(row=row, column=1).value
    staff_name = number_master_sheet.cell(row=row, column=2).value
    NP_number = number_master_sheet.cell(row=row, column=3).value
    temp = {PD_number: (NP_number, staff_name)}
    number_name_dict.update(temp)
# ============================ 写入Master字典 ============================#



if nowtime.day < 10:
    nowMonth = nowtime.month - 1
    if nowMonth == 0:  # 处理跨年的情况
        nowMonth = 12
else:
    nowMonth = nowtime.month
nowMonth = str(nowMonth).zfill(2)
print("処理中:{}月".format(nowMonth))

root = tk.Tk()
root.withdraw()
root.lift()  # 提升窗口
root.attributes('-topmost', True)  # 置顶窗口

pdf_path = filedialog.askopenfilename()
root.destroy()  # 关闭主窗口
print(pdf_path)

# ============================ 辅助函数 ============================#


def write_to_excel(ws, name, data_list):
    """
    立即写入 Excel, 一行データ = (名字, 日期, 上班时间, 下班时间)
    """
    if not data_list:
        print(f"⚠️ 数据列表为空: {name}")
        return  # 没有数据，不写入

    for date, time1, time2 in data_list:
        # 解析日期，提取日(DD)
        match = re.match(r"\d{2}/(\d{2})", date)
        if not match:
            print(f"⚠️ 日期格式错误: {date}")
            continue  # 跳过错误日期

        day = int(match.group(1))
        row_idx = 8 + day  # Excel 行索引，例如 1日 -> 9行, 2日 -> 10行, ..., 31日 -> 39行

        ws[f"A{row_idx}"] = date  # A列 -> 日期

        try:
            tobj = datetime.strptime(time1, "%H:%M").time()

            # 调整出勤时间（如果超过20分，进1到整点）
            if tobj.minute > 20:
                corrected_hour = (tobj.hour + 1) % 24 #你需要在 23:30 以后 的情况下，让小时变为 00，而不是 24。可以这样修改
                """
                示例测试
                原始时间	结果
                08:25	09:00
                22:51	23:00
                23:35	00:00
                23:59	00:00
                """
                corrected_time = f"{corrected_hour}:00"
                corrected_time = datetime.strptime(corrected_time, "%H:%M").time()

            else:
                corrected_time = tobj


            ws[f"C{row_idx}"] = corrected_time  # 写入修正后的时间

            # 休息时间计算
            if corrected_time.hour >= 18 or corrected_time.hour == 0:  # 夜班,0時出勤は夜勤です。
                ws[f"E{row_idx}"] = datetime.strptime("1:00", "%H:%M").time()
                ws[f"F{row_idx}"] = datetime.strptime("2:00", "%H:%M").time()
            elif corrected_time.hour < 12:  # 白班
                ws[f"E{row_idx}"] = datetime.strptime("12:00", "%H:%M").time()
                ws[f"F{row_idx}"] = datetime.strptime("13:00", "%H:%M").time()
            else:
                pass

        except ValueError:
            ws[f"C{row_idx}"] = time1  # 若转换失败，则原样写入

        try:
            time2 = datetime.strptime(time2, "%H:%M").time()
            ws[f"D{row_idx}"] = time2 # 下班时间

        except ValueError:
            ws[f"D{row_idx}"] = time2  # 若转换失败，则原样写入



    ws["D4"] = name  # 将姓名写入 D4
    sheet_title = f"{page_num}_{name}"[:20]
    ws.title = sheet_title
    output_xlsx = "CCMacro_2025_{}.xlsm".format(nowMonth)
    wb.save(output_xlsx)
    print(f"📄 已写入 Excel: {name} - {len(data_list)} 条记录")


def full_width_to_half_width(text: str) -> str:
    """将全角字符转换为半角字符 (ここでは主にコロン「：」から「:」への変換を示しています)。"""
    full_width_chars = "："
    half_width_chars = ":"
    trans = str.maketrans(full_width_chars, half_width_chars)
    return text.translate(trans)


def process_date(text: str) -> str:
    """
    '1日'、'14日' などの日付形式をマッチし、'MM/DD' 形式の日付を返します。
    ここで MM は現在の月（2桁）、DD は抽出した日付（2桁）です。
    """
    date_pattern = r"(\d{1,2})日"  # 1-2桁の数字に後続する '日' をマッチ
    match = re.match(date_pattern, text)
    if match:
        day = match.group(1).zfill(2)  # 日付を2桁にゼロパディング
        return f"{nowMonth}/{day}"     # 'MM/DD' 形式を返す
    return text


def parse_times(text: str) -> list:
    """
    テキストから時間を抽出（複数の時間と単一の時間の両方をサポート）、時間文字列のリスト (0~2 個) を返します：
      - "当日0:24当日1:23" -> ["0:24", "1:23"]
      - "前日21:47当日8:06" -> ["21:47", "8:06"]
      - "当日8:06" -> ["8:06"]
      - "関係ないテキスト" -> []
    """
    text = full_width_to_half_width(text)
    pattern = r'(?:当日|前日)?\d{1,2}:\d{2}'
    items = re.findall(pattern, text)
    if len(items) >= 2:
        t1 = re.sub(r'(当日|前日)', '', items[0])
        t2 = re.sub(r'(当日|前日)', '', items[1])
        return [t1, t2]
    elif len(items) == 1:
        t = re.sub(r'(当日|前日)', '', items[0])
        return [t]
    else:
        return []


def zero_pad_time(time_str):
    """
    "6:06" を "06:06" に、"0:15" を "00:15" に変換します。'H:MM' または 'HH:MM' でない場合は元の文字列を返します。
    """
    pattern = r'^(\d{1,2}):(\d{2})$'
    match = re.match(pattern, time_str)
    if match:
        hour, minute = match.groups()
        hour = hour.zfill(2)
        return f"{hour}:{minute}"
    return time_str


# ============================ 主ロジック ============================#




with fitz.open(pdf_path) as pdf:
    # pdf_pages = pdf.page_count  # 你的 PDF 总页数
    # progress_bar0 = tqdm(range(pdf_pages), desc="ページOCR 处理中", ncols=80, position=0, leave=True)
    # progress_bar1 = tqdm(range(pdf_pages), desc="個人OCR 处理中", ncols=80, position=1, leave=True)

    for page_num in tqdm(range(pdf.page_count), desc="OCR読み取り中・・・", ncols=80,unit="ページ",position=0,leave=True):
        # tqdm.write(f"\n开始 OCR Page {page_num + 1}")
        # tqdm.write(f"📄 已写入 Excel: Page {page_num + 1}")
        # tqdm(range(pdf.page_count), desc="OCR読み取り中・・・", ncols=80, unit="ページ", position=0, leave=True).update()
        page = pdf.load_page(page_num)
        matrix = fitz.Matrix(200 / 72, 200 / 72)
        width, height = page.rect.width, page.rect.height
        crop_rect = fitz.Rect(width * 0.0, height * 0.0, width * 0.6, height * 1)
        page.set_cropbox(crop_rect)

        pix = page.get_pixmap(matrix=matrix)
        img_path = f"output\\PDFToPNG_PAGE_{page_num + 1}.png"
        pix.save(img_path)

        print(f"\n开始OCR Page{page_num + 1}")
        result = ocr.ocr(img_path, cls=False)


        current_line_refY = None

        data_list = []  # 存储所有的 (日期, 时间1, 时间2)
        time_list = []  # 暂存一行中的时间对
        current_date = None  # 当前日期
        last_name = None  # 记录上一个姓名

        for i, line in tqdm(enumerate(result[0]),desc=f"{page_num}ページ内：OCR読み取り中・・・", ncols=80,unit="データ数",position=1,leave=True):
            text = line[1][0]
            coords = line[0]
            center_pointX = sum(pt[0] for pt in coords) / 4
            center_pointY = sum(pt[1] for pt in coords) / 4
            print(f"読み取り中：{text}")

            # **1️⃣ 提取姓名（NP 开头）**
            if "NP" in text:
                name_now = text.strip()

                # 结算上一个人的数据
                if last_name and data_list:
                    print(f"✅ 结算 {last_name} 的数据: {data_list}")
                    ws = wb.copy_worksheet(source)
                    write_to_excel(ws, last_name, data_list)
                    data_list.clear()  # 清空数据

                last_name = name_now  # 更新姓名
                continue  # 继续下一行

            # **2️⃣ 识别日期**
            dtmp = process_date(text)
            print(f"日付認識開始{dtmp}")
            if dtmp != text:
                current_date = dtmp  # 更新日期
                print(f"日付認識OK：{current_date}")
                time_list.clear()  # 清空时间缓存
                continue  # 继续下一行

            # **3️⃣ 识别时间**
            parsed = parse_times(text)
            print(f"時間認識開始{parsed}")
            print("時間数：",len(time_list))
            if parsed:
                print(f"時間認識OK：{parsed}")
                time_list.extend(parsed)  # 加入时间列表

                # ⚠️ 只有两次时间才存入 data_list
                if len(time_list) == 2:
                    data_list.append((current_date, time_list[0], time_list[1]))  # (日期, 时间1, 时间2)
                    time_list.clear()  # 清空，准备下一行
                continue  # 继续下一行

        # **4️⃣ 处理最后一批数据**
        if last_name and data_list:
            print(f"✅ 最后结算 {last_name} 的数据: {data_list}")
            ws = wb.copy_worksheet(source)
            write_to_excel(ws, last_name, data_list)

print("全部处理完成。")