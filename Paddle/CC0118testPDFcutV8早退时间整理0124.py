import re
import fitz
import openpyxl
from paddleocr import PaddleOCR
from datetime import datetime
# ============================ 辅助函数 ============================#
# ============================ 名前処理 ============================#
def process_name_pd2(text, number_name_dict):
    """PD2 の名字処理逻辑"""
    name_now, testnumber, testname = None, "不明", "未识别氏名"

    if "氏名" in text:
        tmp = text.replace(":", '').replace("：", '').replace("氏名", '')
        print("14",tmp)
        name_now = tmp.strip() if tmp else "未识别氏名"
        try:
            testnumber = number_name_dict[name_now][0]
            print(testnumber)
            testname = number_name_dict[name_now][1]
            print(testname)
        except KeyError:
            testnumber = "不明"
            testname = name_now

    return name_now, testnumber, testname, None


def process_name_cc(text, pending_name_part, coords, config, number_name_dict):
    """CC の名字処理逻辑"""
    name_now, testnumber, testname = None, "不明", "未识别姓名"

    name_key = config['name_key']
    THRESHOLD_Y = config['y_threshold']
    center_pointX = sum(pt[0] for pt in coords) / 4
    center_pointY = sum(pt[1] for pt in coords) / 4

    if text.startswith(name_key):
        if pending_name_part:
            # 検查是否是名字の第二部分
            if abs(center_pointX - pending_name_part[1]) < 180 and abs(center_pointY - pending_name_part[2]) < THRESHOLD_Y:
                name_now = pending_name_part[0] + text.replace(name_key, '').strip()
                pending_name_part = None
            else:
                # 現在の段が名前の継続でない場合、現在の名前を保存
                name_now = pending_name_part[0]
                pending_name_part = None
        else:
            # 名前の最初の部分を検出
            pending_name_part = (text.strip(), center_pointX, center_pointY)

    # 名前が連結された場合、testnumber と testname を取得
    if name_now:
        try:
            testnumber = number_name_dict[name_now][0]
            testname = number_name_dict[name_now][1]
        except KeyError:
            testnumber = "不明"
            testname = name_now

    return name_now, testnumber, testname, pending_name_part

# ============================ 名前処理完了 ============================#
factory_configs = {
    "PD2": {
        "crop_rect": [0.0, 0.15, 0.48, 0.75],
        "y_threshold": 5,
        "name_key": "氏名",
        "name_processor": process_name_pd2  # PD2 使用的名字処理関数
    },
    "CC": {
        "crop_rect": [0.1, 0.2, 0.6, 0.9],
        "y_threshold": 2.0,
        "name_key": "NP",
        "name_processor": process_name_cc  # CC 使用的名字処理関数
    }
}
pdf_path = r"pd2.pdf"
current_factory = "PD2"
config = factory_configs[current_factory]
name_processor = factory_configs[current_factory]["name_processor"]




def full_width_to_half_width(text: str) -> str:
    """将全角字符转换为半角字符 (这里主要示范冒号“：”到“:”の変換)。"""
    full_width_chars = "："
    half_width_chars = ":"
    trans = str.maketrans(full_width_chars, half_width_chars)
    return text.translate(trans)


def process_date(text: str) -> str:
    """
    匹配日期格式：
    - PD2: '07/14' 或 '07-14' 格式。
    - CC: 1-2位数字加“日”（例えば '2日', '12日'）。
    返却するマッチした日付部分または原文本。
    """
    # 匹配 PD2 日期フォーマット
    pd2_date_pattern = r"\d{2}[/-]\d{2}"
    # 匹配 CC 日期フォーマット (1-2位数字の後に「日」)
    cc_date_pattern = r"\d{1,2}日"

    # まず PD2 日期を優先的にマッチ
    match = re.match(pd2_date_pattern, text)
    if match:
        return match.group()

    # 次に CC 日期をマッチ
    match = re.match(cc_date_pattern, text)
    if match:
        # 日付部分を抽出 (「日」を削除)
        return match.group().replace("日", "")

    # どちらにもマッチしない場合、原文本を返却
    return text


def parse_times(text: str) -> list:
    """
    テキストから時間を抽出（複数時間のマージと単一時間のサポート）、時間文字列リスト (0~2 個) を返却：
      - "当日0:24当日1:23" -> ["0:24", "1:23"]
      - "前日21:47当日8:06" -> ["21:47", "8:06"]
      - "当日8:06" -> ["8:06"]
      - "関係ないテキスト" -> []
    """
    text = full_width_to_half_width(text)
    pattern = r'(?:当日|前日)?\d{1,2}:\d{2}'
    items = re.findall(pattern, text)
    if len(items) >= 2:
        t1 = re.sub(r'(当日|前日)', '', items[0])
        t2 = re.sub(r'(当日|前日)', '', items[1])
        return [t1, t2]
    elif len(items) == 1:
        t = re.sub(r'(当日|前日)', '', items[0])
        return [t]
    else:
        return []


def zero_pad_time(time_str):
    """
    "6:06" を "06:06" に変換；"0:15" を "00:15" に変換；'H:MM' または 'HH:MM' でない場合は元のまま返却。
    """
    pattern = r'^(\d{1,2}):(\d{2})$'
    match = re.match(pattern, time_str)
    if match:
        hour, minute = match.groups()
        hour = hour.zfill(2)
        return f"{hour}:{minute}"
    return time_str


# 在設定に処理関数をバインド


# ============================ 主論理 ============================#



ocr = PaddleOCR(
    use_angle_cls=False,
    # table=True,
    # table_algorithm='TableAttn',
    # table_max_len=488,
    #  軽量モデル等可自行指定
    lang='ch',
)

wb = openpyxl.load_workbook("PDtestM.xlsm",keep_vba=True)
source = wb["99999　ニッセープロダクツ"]
number_master_sheet = wb['Number_Master']
number_name_dict = {}
for row in range(1,number_master_sheet.max_row+1):
    PD_number = number_master_sheet.cell(row=row,column=1).value
    staff_name = number_master_sheet.cell(row=row,column=2).value
    NP_number = number_master_sheet.cell(row=row,column=3).value
    temp = {PD_number:(NP_number,staff_name)}
    number_name_dict.update(temp)


name_now = None
current_date = None

with fitz.open(pdf_path) as pdf:
    print(f"PDF 页数: {pdf.page_count}")
    for page_num in range(pdf.page_count):
        page_rows_data = []  # 本ページのすべての行のデータ記録 [(name, date, [times...])...]

        # PDF ページを開き、レンダリング
        page = pdf.load_page(page_num)
        print(f"Page {page_num}: size={page.rect.width}x{page.rect.height}")
        matrix = fitz.Matrix(200 / 72, 200 / 72)
        width, height = page.rect.width, page.rect.height


        crop_ratio = config['crop_rect']  # 例えば [0.0, 0.15, 0.48, 0.75]
        crop_rect = fitz.Rect(
            crop_ratio[0] * width,  # 左上角 X
            crop_ratio[1] * height,  # 左上角 Y
            crop_ratio[2] * width,  # 右下角 X
            crop_ratio[3] * height  # 右下角 Y
        )
        page.set_cropbox(crop_rect)


        pix = page.get_pixmap(matrix=matrix)
        img_path = f"output\\PDFToPNG_PAGE_{page_num + 1}.png"
        pix.save(img_path)

        # OCR
        print(f"\n開始OCR Page{page_num + 1}")
        result = ocr.ocr(img_path, cls=False)
        name_now = None
        pending_name_part = None  # 未完成の名前の部分をキャッシュ

        # 現在の行の (time_str, center_x, center_y) をキャッシュ
        current_line_times = []
        # 現在の行の参照y座標を記録し、改行を判断する
        # 初期値は None で、最初の時間があるときに値を設定
        current_line_refY = None
        THRESHOLD_Y = config['y_threshold']  # 許容される上下誤差 CC1.5
        name_key = config['name_key']

        for i, line in enumerate(result[0]):
            text = line[1][0]
            coords = line[0]
            center_pointX = sum(pt[0] for pt in coords) / 4
            center_pointY = sum(pt[1] for pt in coords) / 4

            print(f"Page {page_num + 1} 行 {i}: center=({center_pointX:.1f}, {center_pointY:.1f}), text={text}")

            if current_line_refY is None:
                # これは現在の行の最初のものであることを示す
                current_line_refY = center_pointY

            else:
                # 前の行の Y 座標との差が THRESHOLD_Y より大きい => 改行
                if abs(center_pointY - current_line_refY) > THRESHOLD_Y:
                    # まず前の行を決算
                    if current_line_times:
                        # ソートまたは収集
                        current_line_times.sort(key=lambda x: x[1])  # X でソート
                        # (name_now, current_date, times) を page_rows_data に追加
                        page_rows_data.append((name_now, current_date, current_line_times[:]))
                        print("改行、前の行が決算:", current_line_times)

                        # 現在の行のキャッシュをクリア
                        current_line_times.clear()

                    # 行の参照Yを更新
                    current_line_refY = center_pointY

            # 工場タイプに応じてパラメータを動的に調整
            if current_factory == "PD2":
                name_now, testnumber, testname, pending_name_part = name_processor(
                    text, number_name_dict
                )
            elif current_factory == "CC":
                name_now, testnumber, testname, pending_name_part = name_processor(
                    text, pending_name_part, coords, config, number_name_dict
                )

            # 結果を出力
            if name_now:
                print(f"名前を認識: {name_now}, 工号: {testnumber}, 姓名: {testname}")

            # テキストに日付が含まれている場合
            dtmp = process_date(text)
            if dtmp != text:
                current_date = dtmp  # 抽出した日付は '07/14' または '2'（数字部分）
                print(f"日付を認識: {current_date}")
                continue

            # 出現する可能性のある時間を解析
            parsed = parse_times(text)
            if not parsed:
                continue  # 時間がない場合は続ける

            # 1~2つの時間がある場合、ループで追加
            for tm in parsed:
                # 同じ行にあるかどうかをチェック：current_line_refY が空の場合、最初のもの
                if current_line_refY is None:
                    current_line_refY = center_pointY
                    current_line_times.append((tm, center_pointX, center_pointY))
                else:
                    # 現在の行参照Y座標との差が THRESHOLD_Y より大きい場合、改行と見なす
                    if abs(center_pointY - current_line_refY) > THRESHOLD_Y:
                        # つまり前の行の times は決算すべき
                        # ソート、page_rows_data に記録
                        if current_line_times:
                            page_rows_data.append((name_now, current_date, current_line_times[:]))
                            print("改行を検出、前の行に収集された時間:", current_line_times)

                        # クリア
                        current_line_times.clear()
                        # 新しい行参照Y
                        current_line_refY = center_pointY
                        # 現在のこの時間を新しい行に追加
                        current_line_times.append((tm, center_pointX, center_pointY))
                    else:
                        # まだ同じ行
                        current_line_times.append((tm, center_pointX, center_pointY))

        # ループが終了した後、まだ決算されていない現在の行があれば page_rows_data に追加
        if current_line_times:
            page_rows_data.append((name_now, current_date, current_line_times[:]))
            current_line_times.clear()

        print(f"--- Page {page_num + 1} 収集: {page_rows_data}")

        # =============== 本ページの行データを Excel に書き込むロジック ===============#
        ws = wb.copy_worksheet(source)
        if page_rows_data:
            # 最後のレコードの名前をシートタイトルとして取得
            sheet_title = page_rows_data[-1][0] or "結果"
            ws.title = sheet_title
        else:
            ws.title = "結果"

        # 例：本ページのすべての行を Excel に書き込む
        # 各行には 2 つの時間がある場合もあるが、3 つまたは 4 つもあるため、自分で処理する必要がある
        # 下記はデモ書き方
        for (line_name, line_date, times_list) in page_rows_data:
            # times_list が空の場合、書き込む必要がない時間 => スキップ
            if not times_list:
                continue

            # line_date から日(DD)を抽出する、例えば "12/01" -> day=1
            match = re.match(r"\d{2}[/-](\d{2})", line_date)
            if not match:
                # 日付フォーマットの解析に失敗した場合、このレコードをスキップ
                continue

            day = int(match.group(1))
            row_idx = 8 + day  # 1 -> 9行, 2 -> 10行, ..., 31 -> 39行

            # X 座標でソート
            times_list.sort(key=lambda x: x[1])

            # 時間文字列を抽出し、ゼロパディング
            time_strs = [zero_pad_time(x[0]) for x in times_list]

            # A列 -> 日付
            ws[f"A{row_idx}"] = line_date

            # 2つの時間がある場合 => C,D; 4つの場合 => C,D,E,F; その他は自分で処理
            for idx, tm_str in enumerate(time_strs):
                # 0->C, 1->D, 2->E, 3->F
                col = chr(ord('C') + idx)
                try:
                    tobj = datetime.strptime(tm_str, "%H:%M").time()
                    ws[f"{col}{row_idx}"] = tobj
                except ValueError:
                    # 時間に変換できない場合、元のまま書き込む
                    ws[f"{col}{row_idx}"] = tm_str

        # 異なる page_num ごとに別ファイルを保存するか？要件に応じて
        # ここでは同じ output に保存するのみをデモ
        ws["H4"].value = testnumber
        ws["D4"].value = testname

        output_xlsx = "CCMacro_2025_01.xlsm"
        wb.save(output_xlsx)
        print(f"Page {page_num + 1} がシート={ws.title}, ファイル={output_xlsx} に書き込まれました")

print("すべての処理が完了しました。")