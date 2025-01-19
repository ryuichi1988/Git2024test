import os
import re
import fitz
import openpyxl

from paddleocr import PaddleOCR
from concurrent.futures import ProcessPoolExecutor, as_completed

# ------------------- 全局 & 初始化函数 ------------------- #
ocr_obj = None  # 全局变量，每个进程内会有一份拷贝


def init_ocr():
    """
    多进程启动时，初始化 PaddleOCR。
    这样每个进程只加载一次模型，避免重复开销。
    你也可设置 enable_mkldnn=True, cpu_threads=6 等试验性能。
    """
    global ocr_obj
    if ocr_obj is None:
        ocr_obj = PaddleOCR(
            use_gpu=False,
            enable_mkldnn=True,
            cpu_threads=6,
            lang='ch',
            table=True,
            table_algorithm='TableAttn',
            table_max_len=488,
            show_log=False
        )


# ------------------- 辅助函数 ------------------- #

def full_width_to_half_width(text: str) -> str:
    """将全角字符转换为半角字符 (这里主要示范冒号“：”到“:”的转换)。"""
    full_width_chars = "："
    half_width_chars = ":"
    trans = str.maketrans(full_width_chars, half_width_chars)
    return text.translate(trans)


def process_date(text: str) -> str:
    """
    使用正则匹配日期格式 (如 '07/14' 或 '07-14')，
    如果匹配成功，返回提取到的日期部分；否则返回原文本。
    """
    date_pattern = r"\d{2}[/-]\d{2}"
    match = re.match(date_pattern, text)
    if match:
        return match.group()
    return text


def parse_times(text: str) -> list:
    """
    提取时间，支持合并或单一时间：
      - "当日0:24当日1:23" -> ["0:24", "1:23"]
      - "前日21:47当日8:06" -> ["21:47", "8:06"]
      - "当日8:06" -> ["8:06"]
      - "无关文本" -> []
    """
    text = full_width_to_half_width(text)
    pattern = r'(?:当日|前日)?\d{1,2}:\d{2}'
    items = re.findall(pattern, text)
    if len(items) >= 2:
        t1 = re.sub(r'(当日|前日)', '', items[0])
        t2 = re.sub(r'(当日|前日)', '', items[1])
        return [t1, t2]
    elif len(items) == 1:
        t = re.sub(r'(当日|前日)', '', items[0])
        return [t]
    else:
        return []


def check_y_coord_consistency(times_collected, threshold=5) -> bool:
    """检查 4 个时间的 Y 坐标是否在同一行 (差值 <= threshold)。"""
    if len(times_collected) != 4:
        return False
    ys = [t[2] for t in times_collected]
    return (max(ys) - min(ys)) <= threshold


def zero_pad_time(time_str):
    """
    将 "6:06" 转换为 "06:06"；"0:15" 转换为 "00:15"；
    若非 'H:MM' 或 'HH:MM' 格式，原样返回。
    """
    pattern = r'^(\d{1,2}):(\d{2})$'
    match = re.match(pattern, time_str)
    if match:
        hour, minute = match.groups()
        hour = hour.zfill(2)
        return f"{hour}:{minute}"
    return time_str


# ------------------- 单页处理函数 (并行入口) ------------------- #

def process_page_task(args):
    """
    每个进程都会调用此函数来处理 PDF 某页。
    返回该页解析到的 rows_data (可能有多行) 以及可能的最后姓名。
    """
    global ocr_obj
    page_num, pdf_path = args

    # 打开 PDF 并渲染当前页
    doc = fitz.open(pdf_path)
    page = doc.load_page(page_num)
    matrix = fitz.Matrix(200 / 72, 200 / 72)  # 设置渲染DPI (可自行调整)
    width, height = page.rect.width, page.rect.height
    crop_rect = fitz.Rect(width * 0.0, height * 0.15, width * 0.48, height * 0.75)
    page.set_cropbox(crop_rect)

    pix = page.get_pixmap(matrix=matrix)
    doc.close()

    # 保存临时图片
    img_path = f"temp_{os.getpid()}_{page_num}.png"
    pix.save(img_path)

    # OCR 识别
    result = ocr_obj.ocr(img_path, cls=False)

    # 删除临时文件
    if os.path.exists(img_path):
        os.remove(img_path)

    # 以下是与原本逻辑相同，解析 name, date, times...
    rows_data = []  # [(name, date, t1, t2, t3, t4), ...]
    name_now = None
    current_date = None
    times_buffer = []  # [(time_str, x, y), ...]

    for line in result[0]:
        text = line[1][0]
        coords = line[0]
        center_pointX = sum(pt[0] for pt in coords) / 4
        center_pointY = sum(pt[1] for pt in coords) / 4

        # 1) 判断是否含 "氏名"
        if "氏名" in text:
            # 去除多余字符
            tmp = text.replace(":", '').replace("：", '').replace("氏名", '')
            name_now = tmp.strip() if tmp else "未识别氏名"
            continue

        # 2) 判断是否是日期
        dtmp = process_date(text)
        if dtmp != text:
            current_date = dtmp
            continue

        # 3) 分析时间
        parsed_times = parse_times(text)
        if parsed_times:
            for t in parsed_times:
                times_buffer.append((t, center_pointX, center_pointY))

        # 4) 如果收集到4个时间，就判断是否在同一行
        if len(times_buffer) == 4:
            if not check_y_coord_consistency(times_buffer, threshold=5):
                times_buffer.clear()
                continue

            # 排序并组合成记录
            times_buffer.sort(key=lambda x: x[1])
            t1, t2, t3, t4 = [tb[0] for tb in times_buffer]

            # 若未识别到 name 或 date，给默认值
            if not name_now:
                name_now = "未识别氏名"
            if not current_date:
                current_date = "XX/XX"

            # 补齐时间格式
            t1 = zero_pad_time(t1)
            t2 = zero_pad_time(t2)
            t3 = zero_pad_time(t3)
            t4 = zero_pad_time(t4)

            row_data = (name_now, current_date, t1, t2, t3, t4)
            rows_data.append(row_data)

            # 清空 time 缓存
            times_buffer.clear()

    # 返回 (page_idx, rows_data)
    # 有时还需要返回 name_now，看你怎么在多页合并时处理
    return (page_num, rows_data)


# ------------------- 主逻辑 (多进程并行) ------------------- #

def main(pdf_path):
    import math

    # 打开 PDF 获取页数
    doc = fitz.open(pdf_path)
    total_pages = doc.page_count
    doc.close()

    # 生成 (page_num, pdf_path) 列表
    page_args = [(p, pdf_path) for p in range(total_pages)]

    all_rows_data = []

    # max_workers=6 充分利用 6 核
    with ProcessPoolExecutor(max_workers=6, initializer=init_ocr) as executor:
        futures = {executor.submit(process_page_task, arg): arg[0] for arg in page_args}
        for fut in as_completed(futures):
            page_idx = futures[fut]
            try:
                page_num, rows_data = fut.result()
                all_rows_data.extend(rows_data)  # 直接汇总
            except Exception as e:
                print(f"[错误] 处理第 {page_idx} 页时出现异常: {e}")

    # 返回收集到的所有行数据
    return all_rows_data


# ------------------- 运行并写入 Excel ------------------- #

if __name__ == "__main__":
    pdf_path = r"C:\Users\timaz\Documents\PythonFile\pd2\example3.pdf"
    print("开始多进程OCR处理...")
    rows_data = main(pdf_path)
    print("处理完成，开始写Excel...\n")

    for rd in rows_data:
        print(rd)

    # 假设接下来写入Excel
    wb = openpyxl.load_workbook("PDtest.xlsx")
    source = wb["99999　ニッセープロダクツ"]
    ws = wb.copy_worksheet(source)

    if rows_data:
        sheet_title = rows_data[-1][0]  # 取最后一条的 name
        ws.title = sheet_title
        # 同时也可以在某处 ws["D4"].value = sheet_title
    else:
        ws.title = "结果"

    # 写入 Excel： 与之前逻辑相同
    # 日期 -> A 列；时间 -> C~F 列；行号 = 8 + 日（1 日 -> 第 9 行）
    for row_item in rows_data:
        name, date_str, t1, t2, t3, t4 = row_item
        match = re.match(r"\d{2}[/-](\d{2})", date_str)
        if match:
            day = int(match.group(1))
        else:
            day = 1
        row_idx = 8 + day

        ws[f"A{row_idx}"] = date_str
        ws[f"C{row_idx}"] = t1
        ws[f"D{row_idx}"] = t2
        ws[f"E{row_idx}"] = t3
        ws[f"F{row_idx}"] = t4

    output_xlsx = "output.xlsx"
    wb.save(output_xlsx)
    print(f"Excel 已保存到: {output_xlsx}, 工作表名称: {ws.title}")
