import re
import fitz
import openpyxl
from paddleocr import PaddleOCR
from datetime import datetime


# ============================ 辅助函数 ============================#

def full_width_to_half_width(text: str) -> str:
    """将全角字符转换为半角字符 (这里主要示范冒号“：”到“:”的转换)。"""
    full_width_chars = "："
    half_width_chars = ":"
    trans = str.maketrans(full_width_chars, half_width_chars)
    return text.translate(trans)


def process_date(text: str) -> str:
    """匹配 '07/14' 或 '07-14' 格式，并返回日期部分。否则返回原文本。"""
    date_pattern = r"\d{2}[/-]\d{2}"
    match = re.match(date_pattern, text)
    if match:
        return match.group()
    return text


def parse_times(text: str) -> list:
    """
    从文本中提取时间（支持合并时间和单一时间），返回时间字符串列表 (0~2 个)：
      - "当日0:24当日1:23" -> ["0:24", "1:23"]
      - "前日21:47当日8:06" -> ["21:47", "8:06"]
      - "当日8:06" -> ["8:06"]
      - "无关文本" -> []
    """
    text = full_width_to_half_width(text)
    pattern = r'(?:当日|前日)?\d{1,2}:\d{2}'
    items = re.findall(pattern, text)
    if len(items) >= 2:
        t1 = re.sub(r'(当日|前日)', '', items[0])
        t2 = re.sub(r'(当日|前日)', '', items[1])
        return [t1, t2]
    elif len(items) == 1:
        t = re.sub(r'(当日|前日)', '', items[0])
        return [t]
    else:
        return []


def zero_pad_time(time_str):
    """
    将 "6:06" 转换为 "06:06"；"0:15" 转换为 "00:15"；若非 'H:MM' 或 'HH:MM' 则原样返回。
    """
    pattern = r'^(\d{1,2}):(\d{2})$'
    match = re.match(pattern, time_str)
    if match:
        hour, minute = match.groups()
        hour = hour.zfill(2)
        return f"{hour}:{minute}"
    return time_str


# ============================ 主逻辑 ============================#

pdf_path = r'C:\Users\timaz\Documents\PythonFile\pd2\example2.pdf'

ocr = PaddleOCR(
    use_angle_cls=False,
    # table=True,
    # table_algorithm='TableAttn',
    # table_max_len=488,
    #  轻量模型等可自行指定
    lang='ch',
)

wb = openpyxl.load_workbook("PDtest.xlsx")
source = wb["99999　ニッセープロダクツ"]
number_master_sheet = wb['Number_Master']
number_name_dict = {}
for row in range(1,number_master_sheet.max_row):
    PD_number = number_master_sheet.cell(row=row,column=1).value
    staff_name = number_master_sheet.cell(row=row,column=2).value
    NP_number = number_master_sheet.cell(row=row,column=3).value
    temp = {PD_number:(NP_number,staff_name)}
    number_name_dict.update(temp)


name_now = None
current_date = None

with fitz.open(pdf_path) as pdf:
    for page_num in range(pdf.page_count):
        page_rows_data = []  # 本页所有行的数据记录 [(name, date, [times...])...]

        # 打开并渲染 PDF 页面
        page = pdf.load_page(page_num)
        matrix = fitz.Matrix(200 / 72, 200 / 72)
        width, height = page.rect.width, page.rect.height
        crop_rect = fitz.Rect(width * 0.0, height * 0.15, width * 0.48, height * 0.75)
        page.set_cropbox(crop_rect)

        pix = page.get_pixmap(matrix=matrix)
        img_path = f"output\\PDFToPNG_PAGE_{page_num + 1}.png"
        pix.save(img_path)

        # OCR
        print(f"\n开始OCR Page{page_num + 1}")
        result = ocr.ocr(img_path, cls=False)

        # 缓存当前行的 (time_str, center_x, center_y)
        current_line_times = []
        # 记录当前行的参考y坐标，用于判断是否换行
        # 初始化为 None，待第一条时间来时赋值
        current_line_refY = None
        THRESHOLD_Y = 5  # 允许的上下误差

        for i, line in enumerate(result[0]):
            text = line[1][0]
            coords = line[0]
            center_pointX = sum(pt[0] for pt in coords) / 4
            center_pointY = sum(pt[1] for pt in coords) / 4

            print(f"Page {page_num + 1} line {i}: center=({center_pointX:.1f}, {center_pointY:.1f}), text={text}")

            if current_line_refY is None:
                # 说明这是本行第一条
                current_line_refY = center_pointY

            else:
                # 如果与之前行的 Y 坐标差距 > THRESHOLD_Y => 换行
                if abs(center_pointY - current_line_refY) > THRESHOLD_Y:
                    # 先结算上一行
                    if current_line_times:
                        # 做排序或收集
                        current_line_times.sort(key=lambda x: x[1])  # 按X排序
                        # 将上一行的 (name_now, current_date, times) 加入到 page_rows_data
                        page_rows_data.append((name_now, current_date, current_line_times[:]))
                        print("换行，上一行已结算:", current_line_times)

                        # 清空当前行的缓存
                        current_line_times.clear()

                    # 更新行的参考Y
                    current_line_refY = center_pointY







            # 如果文本包含 "氏名"
            if "氏名" in text:
                tmp = text.replace(":", '').replace("：", '').replace("氏名", '')
                name_now = tmp.strip() if tmp else "未识别氏名"
                try:
                    testnumber = number_name_dict[name_now][0]
                except KeyError:
                    testnumber = 0
                testname = number_name_dict[name_now][1]
                continue

            # 如果文本包含日期
            dtmp = process_date(text)
            if dtmp != text:
                current_date = dtmp
                continue

            # 解析可能出现的时间
            parsed = parse_times(text)
            if not parsed:
                continue  # 没有时间则继续

            # 有1~2个时间，循环加入
            for tm in parsed:
                # 检查是否是同一行：若 current_line_refY 为空，则是第一条
                if current_line_refY is None:
                    current_line_refY = center_pointY
                    current_line_times.append((tm, center_pointX, center_pointY))
                else:
                    # 如果和当前行参考Y坐标差距 > THRESHOLD_Y，认为换行了
                    if abs(center_pointY - current_line_refY) > THRESHOLD_Y:
                        # 说明上一行的 times 应该结算
                        # 进行排序、记录到 page_rows_data
                        if current_line_times:
                            page_rows_data.append((name_now, current_date, current_line_times[:]))
                            print("检测到换行，上一行收集到的时间：", current_line_times)

                        # 清空
                        current_line_times.clear()
                        # 新的行参考Y
                        current_line_refY = center_pointY
                        # 将当前这条时间放到新行
                        current_line_times.append((tm, center_pointX, center_pointY))
                    else:
                        # 还是同一行
                        current_line_times.append((tm, center_pointX, center_pointY))

        # 循环结束后，如果还有未结算的当前行，也放进 page_rows_data
        if current_line_times:
            page_rows_data.append((name_now, current_date, current_line_times[:]))
            current_line_times.clear()

        print(f"--- Page {page_num + 1} 收集到: {page_rows_data}")

        # =============== 将本页的行数据写入 Excel 逻辑 ===============#
        ws = wb.copy_worksheet(source)
        if page_rows_data:
            # 取最后一条记录中的姓名作为sheet标题
            sheet_title = page_rows_data[-1][0] or "结果"
            ws.title = sheet_title
        else:
            ws.title = "结果"

        # 示例：把本页所有行写入Excel
        # 每行可能 2 个时间，也可能 3 或 4, 需自行处理
        # 下面只是示范写法
        for (line_name, line_date, times_list) in page_rows_data:
            # 如果 times_list 为空，就没有需要写入的时间 => 跳过
            if not times_list:
                continue

            # 从 line_date 中提取日(DD)，例如 "12/01" -> day=1
            match = re.match(r"\d{2}[/-](\d{2})", line_date)
            if not match:
                # 如果日期格式解析失败，就跳过本条
                continue

            day = int(match.group(1))
            row_idx = 8 + day  # 1 -> 9行, 2 -> 10行, ..., 31 -> 39行

            # 按 X 坐标排序
            times_list.sort(key=lambda x: x[1])

            # 提取并补零时间字符串
            time_strs = [zero_pad_time(x[0]) for x in times_list]

            # A列 -> 日期
            ws[f"A{row_idx}"] = line_date

            # 如果只有2个时间 => 写 C,D; 如果4个 => 写 C,D,E,F; 其余自行处理
            for idx, tm_str in enumerate(time_strs):
                # 0->C, 1->D, 2->E, 3->F
                col = chr(ord('C') + idx)
                try:
                    tobj = datetime.strptime(tm_str, "%H:%M").time()
                    ws[f"{col}{row_idx}"] = tobj
                except ValueError:
                    # 若转换为时间失败，就原样写入
                    ws[f"{col}{row_idx}"] = tm_str

        # 不同 page_num 要不要另存一个文件？按你需求
        # 这里只演示存为同一个 output
        ws["H4"].value = testnumber
        ws["D4"].value = testname

        output_xlsx = "output.xlsx"
        wb.save(output_xlsx)
        print(f"Page {page_num + 1} 已写入 Sheet={ws.title}, 文件={output_xlsx}")

print("全部处理完成。")
