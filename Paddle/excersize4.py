import numpy as np
import re
import fitz
import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import subprocess
from datetime import datetime,timedelta
from openpyxl.utils import get_column_letter  # 新增导入


#　TKデフォルト、ファイルを開く。
root = tk.Tk()
root.withdraw()
root.lift()  # 提升窗口
root.attributes('-topmost', True)  # 置顶窗口

pdf_path = filedialog.askopenfilename(title="PDFデータをお選びください。",filetypes=[("pdfファイル","*.pdf")],defaultextension=".pdf")
root.destroy()  # 关闭主窗口
print(pdf_path)

# **🔹 合并相同名字的出勤记录**
merged_data = {}

Person_Sum_Time_List = []

output_xlsx = "CCMacro_2025_02test.xlsm"
# 在现有代码的 structured_array = ... 之后添加以下内容

# 创建日期匹配正则（优化版）
date_pattern = re.compile(r"(\d{1,2})日")  # 匹配 "x日" 格式



# 结果存储
structured_data = []
current_group = None

#時間now
now = datetime.now()
#　作業時間処理（前月）
opration_date = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
print(opration_date)  # 输出：2025-01-01 16:23:00

#　エラーメッセージリスト（txt用）
final_mention_list = []

#　ファイル
wb = openpyxl.load_workbook("CCtestM.xlsm", keep_vba=True)
source = wb["99999　ニッセープロダクツ"]
number_master_sheet = wb['Number_Master']
number_name_dict = {}
name_now = None
last_name = None  # 记录名字

#　NP番号DICT処理　（MASTER）
for row in range(1, number_master_sheet.max_row + 1):
    CCNP_number = number_master_sheet.cell(row=row, column=1).value
    staff_name = number_master_sheet.cell(row=row, column=2).value
    #NP_number = CCNP_number
    temp = {CCNP_number:staff_name}
    number_name_dict.update(temp)





    """
    要实现将包含全角冒号或其它非标准格式的时间字符串（如8：30）转换为datetime.timedelta类型，可以按照以下方案实现：
    """
def custom_timedelta_parser(time_str: str) -> timedelta:
    """支持全角冒号和紧凑格式的时间转换"""
    # 统一符号和格式处理
    clean_str = time_str.replace("：", ":")  # 替换全角冒号
    if ':' not in clean_str:  # 处理类似"0830"的格式
        if len(clean_str) < 3:
            clean_str = f"{clean_str.zfill(2)}:00"  # "8"→"08:00"
        else:
            clean_str = f"{clean_str[:-2].zfill(2)}:{clean_str[-2:]}"  # "830"→"08:30"

    # 分割并转换为timedelta
    parts = clean_str.split(':')
    if len(parts) != 2:
        raise ValueError(f"无效时间格式: {time_str}")

    return timedelta(
        hours=int(parts[0]),
        minutes=int(parts[1])
    )


"""
根据您的需求，我们可以编写一个函数来实现这个功能。
这个函数将接受name作为参数，遍历字典，查找匹配的staff_name，
并返回相应的NP_number。如果找不到匹配的名字，
则返回"不明"。以下是实现这个功能的代码：

使用示例
name_to_search = "张三"  # 要查找的名字
result = find_NP_number(name_to_search, number_name_dict)
print(f"{name_to_search}的NP号码是：{result}")
"""
def find_NP_number(name, number_name_dict):
    for CCNP_number, staff_name in number_name_dict.items():
        if staff_name == name:
            return CCNP_number
    return "不明"


#　HHMM処理
def delta_to_hhmm(delta):
    if delta.days < 0:
        return "00:00"
    total_hours = delta.days * 24 + delta.seconds // 3600
    minutes = (delta.seconds % 3600) // 60
    return f"{total_hours:02d}:{minutes:02d}"

#　退勤時間10分延長
def adjust_off_time(time_str):
    """退勤时间增加10分钟"""
    try:
        # 统一格式处理（兼容 "1700" 和 "17:00" 格式）
        clean_time = time_str.replace("：", ":").strip()
        if ':' not in clean_time and len(clean_time) == 4:
            clean_time = f"{clean_time[:2]}:{clean_time[2:]}"

        # 转换为时间对象
        time_obj = datetime.strptime(clean_time, "%H:%M")

        # 增加10分钟
        adjusted = time_obj + timedelta(minutes=10)
        return adjusted.strftime("%H:%M")

    except ValueError as e:
        # 记录格式错误
        final_mention_list.append((
            name,
            np_number,
            record_group,
            f"退勤时间格式异常: {time_str} (错误: {str(e)})"
        ))
        return time_str  # 保持原始值


#　0SHEET書き込み作業
def write_0sheet(wb):
    ws = wb["0SHEET"]
    sheet0_d1_value = opration_date.strftime("%Y年%m月")
    ws["D1"].value = sheet0_d1_value
    # if len(structured_array) < 200:
    #     for temp_row in range(len(structured_array) ,203 , 1):
    #         ws.row_dimensions[temp_row].hidden = True
    # else:
    #     final_mention_list.append("0sheetの行が足りない。ｽﾀｯﾌ数200超えています？？")


    # 修正列字母生成方式
    date_column_map = {day: get_column_letter(4 + day) for day in range(1, 32)}  # E=5对应day=1

    for row_idx, (name, np_number, records) in enumerate(structured_array, start=4):
        ws[f'C{row_idx}'] = np_number
        ws[f'D{row_idx}'] = name
        ws[f'AN{row_idx}'] = Person_Sum_Time_List[row_idx-4]
"""
        daily_hours = {}

        for record in records:
            try:
                # 添加数组越界检查
                if len(record) < 4:
                    raise ValueError("记录格式不完整")

                day_str = re.search(r'(\d{1,2})日', record[0]).group(1)
                day = int(day_str)

                # 时间格式处理
                # 修改后的时间计算部分
                try:
                    start = datetime.strptime(record[2].strip(), "%H:%M")
                    end = datetime.strptime(record[3].strip(), "%H:%M")

                    # 处理跨日
                    if end < start:
                        end += timedelta(days=1)

                    # 计算总时间并扣除休息
                    total_delta = end - start
                    if total_delta > timedelta(hours=3):
                        work_delta = total_delta - timedelta(hours=1)  # 关键修改点
                    else:
                        work_delta = total_delta
                        final_mention_list.append([name,np_number,record[0],work_delta,"実働少ないです。手動で確認お願いします。"])
                        pass

                    # 处理负数工作时间
                    if work_delta.total_seconds() < 0:
                        raise ValueError(f"工作时间不足1小时（总时长：{total_delta}）")

                    daily_hours[day] = delta_to_hhmm(work_delta)

                except ValueError as e:
                    error_msg = f"工时计算错误: {str(e)}"
                    final_mention_list.append((name, np_number, record, error_msg))
                    daily_hours[day] = 0  # 按0小时记录或保持为负值

                # print(start)
                # print(end)
                # print(work_delta)
                # print(daily_hours[day])

            except Exception as e:
                error_msg = f"记录处理失败: {str(e)}"
                final_mention_list.append((name, np_number, str(record), error_msg))
                continue

        # 写入修正后逻辑
        for day in range(1, 32):
            col = date_column_map.get(day)
            if not col:
                continue

            if day in daily_hours:
                temp_time = datetime.strptime(daily_hours[day], "%H:%M").time()
                ws.cell(
                    row=row_idx,
                    column=4 + day,  # 直接使用列号更可靠
                    value=temp_time
                )

"""

# 读取 PDF 并提取文本
with fitz.open(pdf_path) as doc:
    data = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")
        data.extend(text.split("\n"))  # 按行拆分文本

for item in data:
    if isinstance(item, str) and item.startswith("NP"):
        # 遇到新的 "NPxxx" 开头的字符串，开始新分组
        if current_group:
            structured_data.append(current_group)  # 存储上一个组
        current_group = [item, []]  # 初始化新组
    else:
        # 把数据加入当前组
        if current_group:
            current_group[1].append(item)

# 添加最后一个组
if current_group:
    structured_data.append(current_group)



for name, records in structured_data:
    # print(name,records)

    # **🔴 如果 "name" 以 "合計" 结尾，则跳过**
    if name.endswith("合計"):
        Person_Sum_Time = records[3]  # 個人合計の値をLISTに保存。
        Person_Sum_Time_List.append(Person_Sum_Time)  # 個人合計の値をLISTに保存。
        print(f"Skipping: {name}")  # 调试输出
        continue

    raw_records = np.array(records, dtype=object)  # 转换为 NumPy 数组

    # **按 "x日" 进行分组**
    reshaped_records = []
    temp_group = []

    for item in raw_records:
        if re.match(r"\d{1,2}日", item):  # 如果是 "x日"，表示新的一组
            if temp_group:  # 存入上一个分组
                reshaped_records.append(temp_group[:4])  # 只取前 4 个
            temp_group = [item]  # 开启新分组
        else:
            temp_group.append(item)

    if temp_group:
        reshaped_records.append(temp_group[:4])  # 处理最后一组

    # **🔴 过滤掉第一个元素不包含 "日" 的行**
    filtered_records = [row for row in reshaped_records if row[0].endswith("日")]

    # **如果过滤后数据为空，则跳过这个组**
    if not filtered_records:
        print(f"Skipping group {name} due to no valid records")
        continue

    # 重要ファンクション：出勤、退勤時間変換。
    for record in filtered_records:
        NP__number = find_NP_number(name, number_name_dict)
        work_start_hour = record[1][:2]  # 时间段开始时间 (前 2 位小时)
        work_start_time = record[1][:4]  # 时间段开始时间 (前 2 位小时)
        shift_start_time = custom_timedelta_parser(work_start_time)
        arrival_time = record[2]
        temarrival_time = custom_timedelta_parser(arrival_time)
        try:
            if ":" in work_start_hour:
                raise ValueError
            else:
                pass
            if shift_start_time == timedelta(hours=0) and temarrival_time > timedelta(hours=19):
                record[2] = (datetime.min + shift_start_time).strftime("%H:%M")
            elif shift_start_time > temarrival_time:
                record[2] = (datetime.min + shift_start_time).strftime("%H:%M")
            else:
                pass
        except ValueError:
            final_mention_list.append((name,NP__number,record,"勤務区分エラーです。手動で確認し、出勤時間を修正してください。"))





        # 作業中
        """
        
        
        
        


        arrival_time = record[2]
        arrival_hour, arrival_minute = map(int, record[2].split(":"))  # 上班时间
        new_hour = arrival_hour
        new_minute = arrival_minute
        # print(name)
        # print(name_now)
        # print(name,"start hour", work_start_hour)
        # print(name,"arrival hour", arrival_hour)
        # #print(number_name_dict)
        # print(name)
        NP__number = find_NP_number(name,number_name_dict)

        # **如果上班时间早于时间段开始时间**
        # print("判断開始")
        try:
            if work_start_hour == "00" and arrival_hour > 19:
                # print("判断開始261yes",work_start_hour,arrival_hour)

                arrival_hour = -1

            if arrival_hour < int(work_start_hour):
                # print("判断開始266yes",arrival_hour,work_start_hour)
                new_hour = int(work_start_hour)
                new_minute = 0
                arrival_hour = None
                work_start_hour = None
        except ValueError:
            final_mention_list.append((name,NP__number,record,"勤務区分エラーです。手動で確認し、出勤時間を修正してください。"))


            # **如果是 23 点，改为 00:00**

        # print(name,new_hour,new_minute)
            # **修改上班时间**
        record[2] = f"{new_hour:02d}:{new_minute:02d}"
    """

    # **合并相同名字的记录**
    if name in merged_data:
        #print("**合并相同名字的记录**291 if",name,merged_data)
        merged_data[name] = np.vstack((merged_data[name], filtered_records))
    else:
        #print("**合并相同名字的记录**294 else",name,merged_data)
        merged_data[name] = np.array(filtered_records, dtype=object)

# **转换为 NumPy 数组**
structured_array = np.array([
    (name, find_NP_number(name, number_name_dict), records)  # 实时查询编号
    for name, records in merged_data.items()
], dtype=object)

# **打印结果**
print(structured_array)
list_len = len(structured_array)
"""
n=2
for i in structured_array:
    arr_name_now = i[0]
    number_master_sheet.cell(row=n,column=3).value = arr_name_now
    n += 1
"""

# 时间补零函数
def zero_pad_time(time_str):
    """将 9:5 格式补零为 09:05"""
    if re.match(r"\d{1,2}:\d{1,2}", time_str):
        parts = list(map(int, time_str.split(":")))
        return f"{parts[0]:02d}:{parts[1]:02d}"
    return time_str


# 遍历处理好的数据
for idx, (name, np_number, records) in enumerate(structured_array, start=1):
    try:
        # 复制模板工作表
        new_ws = wb.copy_worksheet(source)

        # 设置工作表名称（NP号码_姓名）
        sheet_title = f"{idx}_{np_number}_{name}"[:31]  # Excel限制31字符
        new_ws.title = sheet_title

        # 写入固定单元格
        new_ws["H4"] = np_number  # 员工号码
        new_ws["D4"] = name  # 员工姓名

        # 处理考勤记录
        for record_group in records:
            # 提取日期（格式：x日）
            original_off = record_group[3]
            record_group[3] = adjust_off_time(original_off)  # 关键修改点

            date_str = record_group[0]
            day_match = date_pattern.match(date_str)
            if not day_match:
                final_mention_list.append((name, np_number, record_group, "日期格式错误"))
                continue

            # 计算目标行号
            day = int(day_match.group(1))
            target_row = 8 + day  # 日期行对应关系

            # 写入日期列
            new_ws[f"A{target_row}"] = f"{opration_date.month}/{day:02d}"  # 假设opration_date已定义

            # 写入时间列
            # 修改时间数据提取方式（取索引2和3）
            time_data = [record_group[2], record_group[3]]  # 索引2:出勤时间，3:退勤时间

            # 列处理逻辑优化（只需要处理C、D两列）
            for col_offset, time_str in enumerate(time_data, start=0):  # 从C列开始偏移
                col_letter = chr(ord('C') + col_offset)  # 0->C,1->D
                try:
                    # 处理时间格式（兼容 "0900" 和 "09:00" 格式）
                    normalized_time = time_str.replace("：", ":")  # 处理全角冒号

                    # 自动补全冒号
                    if ':' not in normalized_time and len(normalized_time) == 4:
                        normalized_time = f"{normalized_time[:2]}:{normalized_time[2:]}"

                    # 转换时间对象
                    time_obj = datetime.strptime(zero_pad_time(normalized_time), "%H:%M").time()
                    new_ws[f"{col_letter}{target_row}"] = time_obj
                except ValueError as e:
                    # 记录原始错误数据
                    error_msg = f"无效时间格式: {time_str} ({str(e)})"
                    new_ws[f"{col_letter}{target_row}"] = time_str
                    final_mention_list.append((
                        name,
                        np_number,
                        str(record_group),  # 记录完整原始数据
                        error_msg
                    ))

    except Exception as e:
        error_msg = f"工作表创建失败: {str(e)}"
        final_mention_list.append((name, np_number, "", error_msg))
        continue

# 最后保存工作簿（保留原有保存逻辑）
write_0sheet(wb)
wb.save(output_xlsx)

if final_mention_list:
    with open(f"CozyRecordMention.txt", "w", encoding="utf-8") as f:
        for sublist in final_mention_list:
            f.write(" ".join(map(str, sublist)) + "\n")  # 用空格分隔元素，并换行
    # 同时打开两个文件（无需等待）
    subprocess.Popen(['notepad.exe', "CozyRecordMention.txt"])



subprocess.Popen(['start', 'excel.exe', output_xlsx], shell=True)