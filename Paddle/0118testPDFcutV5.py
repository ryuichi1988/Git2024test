import re
import fitz
import openpyxl
from paddleocr import PaddleOCR
from datetime import datetime

# ============================ 辅助函数 ============================#

def full_width_to_half_width(text: str) -> str:
    """
    将全角字符转换为半角字符 (这里主要示范冒号“：”到“:”的转换)。
    可根据需要增加其他字符映射。
    """
    full_width_chars = "："
    half_width_chars = ":"
    trans = str.maketrans(full_width_chars, half_width_chars)
    return text.translate(trans)


def process_date(text: str) -> str:
    """
    使用正则匹配日期格式 (如 '07/14' 或 '07-14')，
    如果匹配成功，返回提取到的日期部分；否则返回原文本。
    例如:
      - "07/14日休日" -> "07/14"
      - "07-14木"   -> "07-14"
      - "休息"      -> "休息"
    """
    date_pattern = r"\d{2}[/-]\d{2}"
    match = re.match(date_pattern, text)
    if match:
        return match.group()
    return text


def parse_times(text: str) -> list:
    """
    从文本中提取时间（支持合并时间和单一时间），返回时间字符串列表 (0~2 个)：
      - "当日0:24当日1:23" -> ["0:24", "1:23"]
      - "前日21:47当日8:06" -> ["21:47", "8:06"]
      - "当日8:06" -> ["8:06"]
      - "无关文本" -> []

    核心逻辑：
      1) 全角转半角
      2) 用正则 findall 匹配 "(当日|前日)?\d{1,2}:\d{2}"
      3) 若匹配到 >=2 个，仅取前2个做合并时间
      4) 若只匹配到1个，则单一时间
      5) 去掉“当日”或“前日”后返回
    """
    # 全角 -> 半角
    text = full_width_to_half_width(text)

    # 匹配所有 (当日|前日)? + hh:mm
    pattern = r'(?:当日|前日)?\d{1,2}:\d{2}'
    items = re.findall(pattern, text)

    if len(items) >= 2:
        # 取前两个作为两段时间
        t1 = re.sub(r'(当日|前日)', '', items[0])
        t2 = re.sub(r'(当日|前日)', '', items[1])
        return [t1, t2]
    elif len(items) == 1:
        # 单一时间
        t = re.sub(r'(当日|前日)', '', items[0])
        return [t]
    else:
        return []


def check_y_coord_consistency(times_collected, threshold=5) -> bool:
    """
    检查收集到的 4 个时间是否在同一行 (Y 坐标差 <= threshold)。
    times_collected: [(time_str, center_x, center_y), ...]
    """
    if len(times_collected) != 4:
        return False
    ys = [t[2] for t in times_collected]
    return (max(ys) - min(ys)) <= threshold


def zero_pad_time(time_str):
    """
    将 "6:06" 转换为 "06:06"；"0:15" 转换为 "00:15"；
    对正确的 "HH:MM" 格式进行小时的零补齐，分钟部分若原本为两位数字则不变。
    如果不是 "H:MM" 或 "HH:MM" 格式，则原样返回。
    """
    pattern = r'^(\d{1,2}):(\d{2})$'
    match = re.match(pattern, time_str)
    if match:
        hour, minute = match.groups()
        # 小时补零到两位
        hour = hour.zfill(2)
        # minute 部分已是两位数字，所以保持不变
        return f"{hour}:{minute}"
    return time_str

# ============================ 主逻辑：OCR + 数据收集 ============================#

pdf_path = r'C:\Users\timaz\Documents\PythonFile\pd2\example2.pdf'

# 初始化 OCR（仅示例，实际可用 GPU/CPU, 自行调整）
ocr = PaddleOCR(
    use_angle_cls=False,
    lang='ch',
    table=True,
    table_algorithm='TableAttn',
    table_max_len=488,
)


name_now = None
current_date = None
times_buffer = []  # 用于存放同一行收集到的时间信息 [(time_str, x, y), ...]

wb = openpyxl.load_workbook("PDtest.xlsx")
source = wb["99999　ニッセープロダクツ"]


with fitz.open(pdf_path) as pdf:
    for page_num in range(pdf.page_count):
        rows_data = []  # 清空列表，收集所有记录 (name, date, t1, t2, t3, t4)
        page = pdf.load_page(page_num)

        # 简单的 DPI 设置和裁剪示例
        matrix = fitz.Matrix(200 / 72, 200 / 72)  # ~DPI設置
        width, height = page.rect.width, page.rect.height
        crop_rect = fitz.Rect(width * 0.0, height * 0.15, width * 0.48, height * 0.75)
        page.set_cropbox(crop_rect)

        pix = page.get_pixmap(matrix=matrix)
        img_path = "output\\PDFToPNG_PAGE_{}.png".format(page_num + 1)
        pix.save(img_path)

        # OCR 识别
        print("\n开始OCR{}\n".format(page_num+1))
        result = ocr.ocr(img_path, cls=False)

        n = 0
        for line in result[0]:
            text = line[1][0]
            coords = line[0]
            center_pointX = sum(pt[0] for pt in coords) / 4
            center_pointY = sum(pt[1] for pt in coords) / 4

            print(f"page:{page_num + 1}line:{n}中心({center_pointX:.1f}, {center_pointY:.1f}): {text}")
            n += 1
            # 1) 判断是否含 "氏名"
            if "氏名" in text:
                text = text.replace(":", '')
                text = text.replace("：", '')
                text = text.replace("氏名", '')
                name_now = text  # 或做进一步截取提取真实姓名
                print("氏名:", name_now)
                continue

            # 2) 判断是否为日期
            dtmp = process_date(text)
            if dtmp != text:
                current_date = dtmp
                print("处理后的日期:", current_date)
                continue

            # 3) 分析时间(合并或单一)
            parsed_times = parse_times(text)
            if parsed_times:
                for t in parsed_times:
                    times_buffer.append((t, center_pointX, center_pointY))
                    print("收集到时间 ->", t)

            # 如果收集到4个时间，就判断是否在同一行
            if len(times_buffer) == 4:
                # 检查 Y 坐标一致性
                if not check_y_coord_consistency(times_buffer, threshold=5):
                    print("【错误】Y 坐标不在同一行，放弃这 4 个时间。")
                    times_buffer.clear()
                    continue

                # 一旦确认无误，就排序并组成记录
                times_buffer.sort(key=lambda x: x[1])  # 按 X 坐标排序
                t1, t2, t3, t4 = [tb[0] for tb in times_buffer]

                # 若未识别到 name 或 date，给个默认值
                if not name_now:
                    name_now = "未识别氏名"
                if not current_date:
                    current_date = "XX/XX"

                t1 = zero_pad_time(t1)
                t2 = zero_pad_time(t2)
                t3 = zero_pad_time(t3)
                t4 = zero_pad_time(t4)
                row_data = (name_now, current_date, t1, t2, t3, t4)
                rows_data.append(row_data)
                print("收集到一条完整记录:", row_data)

                times_buffer.clear()


        print(f"\n===== 最终收集到的page {page_num+1} =====")
        for rd in rows_data:
            print(rd)
        ws = wb.copy_worksheet(source)

        # ============================ 写入 Excel 示例 ============================#




        # Sheet 名称：取收集到的最后一个姓名，如果没有则用“结果”
        if rows_data:
            sheet_title = rows_data[-1][0]  # (name, date, t1, t2, t3, t4)[0] -> name
            ws.title = sheet_title
        else:
            ws.title = "结果"

        ws["D4"].value = name_now

        # 将 rows_data 写入 Excel：
        # 规则：日期 -> A 列；时间 -> C~F 列；行号 = 8 + 日（1 日 -> 第 9 行）
        for row_item in rows_data:
            name, date_str, t1, t2, t3, t4 = row_item

            # 从日期中提取日 (DD)，若失败则默认写到第 9 行
            match = re.match(r"\d{2}[/-](\d{2})", date_str)
            if match:
                day = int(match.group(1))  # "07/14" -> day=14
            else:
                day = 1

            row_idx = 8 + day  # 1日 -> 9行；31日 -> 39行

            # A 列写日期
            ws[f"A{row_idx}"] = date_str

            # C~F 列写 4 个时间
            ws[f"C{row_idx}"] = datetime.strptime(t1, "%H:%M").time()
            ws[f"D{row_idx}"] = datetime.strptime(t2, "%H:%M").time()
            ws[f"E{row_idx}"] = datetime.strptime(t3, "%H:%M").time()
            ws[f"F{row_idx}"] = datetime.strptime(t4, "%H:%M").time()

        output_xlsx = "output.xlsx"
        wb.save(output_xlsx)
        print(f"\nExcel 已保存到: {output_xlsx},page: {page_num + 1}  工作表名称: {ws.title}\n")
