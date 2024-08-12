import openpyxl
from datetime import timedelta, datetime, time, date
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from os import system, _exit
from tkinter import messagebox

ListRange = 119
# 作業範囲指定

font = Font(size=18)

fill = PatternFill(patternType='solid', fgColor='E05194')
fill2 = PatternFill(patternType='solid', fgColor='F07A7A')

today = date.today()
thisMonth = today.month  # 今月
todayDay = today.day  # 今日
tomorrowDay = today.day + 1

path = r"C:\1\py\小田原NEWマスタ{}月奥山ver.xlsx".format(str(thisMonth).zfill(2))  # 2桁に指定する

pathtodayfile = r"C:\Users\tkado\Desktop\DC folder\ニッセープロダクツ出勤簿{}月{}日.xlsx".format((str(thisMonth)), str(todayDay))



pathtomorrowfile = r"C:\Users\tkado\Desktop\DC folder\ニッセープロダクツ出勤簿{}月{}日.xlsx".format((str(thisMonth)),
                                                                                      str(tomorrowDay))

filename = ("test{}{}.xlsx".format(thisMonth, todayDay))


def openfile():
    file = openpyxl.load_workbook(path)
    sheet = file.active
    sheet.delete_cols(todayDay + 4, 100)
    for row in range(4 ,ListRange + 1):  # 範囲指定
        # この行ならスキップ：
        

        print("【行：{}】".format(row))

        number = sheet.cell(row=row, column=3).value
        print("【番号：{}】".format(number))

        name = sheet.cell(row=row, column=4).value
        # 　週の１～６　計算。　今日は週７回目と考え
        print("【名前：{}】".format(name))

        day1 = sheet.cell(row=row, column=todayDay - 2).value
        if day1 is None:
            day1 = timedelta(hours=0, minutes=0)
        elif type(day1) == time:
            day1 = timedelta(hours=day1.hour, minutes=day1.minute)
        day2 = sheet.cell(row=row, column=todayDay - 1).value
        if day2 == None:
            day2 = timedelta(hours=0, minutes=0)
        elif type(day2) == time:
            day2 = timedelta(hours=day2.hour, minutes=day2.minute)
        day3 = sheet.cell(row=row, column=todayDay + 0).value
        if day3 == None:
            day3 = timedelta(hours=0, minutes=0)
        elif type(day3) == time:
            day3 = timedelta(hours=day3.hour, minutes=day3.minute)
        day4 = sheet.cell(row=row, column=todayDay + 1).value
        if day4 == None:
            day4 = timedelta(hours=0, minutes=0)
        elif type(day4) == time:
            day4 = timedelta(hours=day4.hour, minutes=day4.minute)
        day5 = sheet.cell(row=row, column=todayDay + 2).value
        if day5 == None:
            day5 = timedelta(hours=0, minutes=0)
        elif type(day5) == time:
            day5 = timedelta(hours=day5.hour, minutes=day5.minute)
        day6 = sheet.cell(row=row, column=todayDay + 3).value
        if day6 == None:
            day6 = timedelta(hours=0, minutes=0)
        elif type(day6) == time:
            day6 = timedelta(hours=day6.hour, minutes=day6.minute)

        day0 = sheet.cell(row=row, column=todayDay - 3).value
        if day0 == None:
            day0 = timedelta(hours=0, minutes=0)
        elif type(day0) == time:
            day0 = timedelta(hours=day0.hour, minutes=day0.minute)

        sixdaysworktime = day1 + day2 + day3 + day4 + day5 + day6

        cellAllCordinateFirst = sheet.cell(row=row,column=5).coordinate
        cellAllCordinateEnd = sheet.cell(row=row,column=todayDay + 3).coordinate

        allTime = "=sum({}:{})".format(cellAllCordinateFirst,cellAllCordinateEnd)
        print("【週6時間：{}】".format(sixdaysworktime))
        print("")



        sheet.cell(row=row, column=todayDay + 4, value=allTime)
        for _ in range(-2, 4):
            sheet.cell(row=row, column=todayDay + _).font = openpyxl.styles.fonts.Font(color='0F00FF')
        sheet.cell(row=row, column=todayDay + 4).number_format = '[hh]:mm'

        sheet.cell(row=row, column=todayDay + 5, value=sixdaysworktime)
        sheet.cell(row=row, column=todayDay + 5).number_format = '[hh]:mm'
        print(number)
        if number is None:
            number=70000000

        else:
            pass

        sheet.cell(row=row, column=todayDay + 6, value=timedelta(hours=28, minutes=0) - sixdaysworktime)
        sheet.cell(row=row, column=todayDay + 6).number_format = '[hh]:mm'

    sheet.column_dimensions[sheet.cell(row=row, column=todayDay + 4).column_letter].width = 20
    sheet.cell(row=3, column=todayDay + 4, value="時間")
    sheet.cell(row=2, column=todayDay + 4, value="{}月合計".format(thisMonth))

    sheet.column_dimensions[sheet.cell(row=row, column=todayDay + 5).column_letter].width = 20
    sheet.cell(row=3, column=todayDay + 5, value="週6日集計")
    sheet.cell(row=2, column=todayDay + 5, value="{}月{}日まで6日".format(thisMonth,todayDay-1))

    sheet.column_dimensions[sheet.cell(row=row, column=todayDay + 6).column_letter].width = 20
    sheet.cell(row=3, column=todayDay + 6, value="28H残り：")
    sheet.cell(row=3, column=todayDay + 6).fill = fill2
    sheet.cell(row=2, column=todayDay + 6, value="28H残時間")
    sheet.cell(row=2, column=todayDay + 6).fill = fill2

    filtercolumn = sheet.cell(row=2, column=todayDay + 6).column_letter
    sheet.auto_filter.ref = "A3:{}{}".format(filtercolumn, ListRange)
    file.save(filename)
    file.close()


if __name__ == "__main__":
    openfile()
    system(filename)
