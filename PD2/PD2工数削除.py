import openpyxl

path1 = r"C:\1\PD2\PD2工数.xlsx"
path2 = r"C:\1\PD2\プライム第二出勤者名簿.xlsx"

file = openpyxl.load_workbook(path1)
file2 = openpyxl.load_workbook(path2)
sheet0 = file["0SHEET"]
sheet2 = file["2在籍者名簿Master"]

NPdate = input("作業の日付を入れてください。")
date01 = NPdate[:2]
date02 = NPdate[2:]

NPnumber = input("{}月{}日のデータ：削除したいｽﾀｯﾌ番号を入力してください。".format(date01,date02))

for cell in sheet0["C"]:
    if cell.value == int(NPnumber):
        sheet0.cell(row=cell.row,column=int(date02)+4).value = None
        break
    else:
        pass

list = file.sheetnames
for i in list:
    # もし番号がシート名に含まれている場合
    if str(NPnumber) in i:
        # シート名
        # print(i)
        PD2kosuExcel_PersonalSheet = file[i]
        # 特定したシートのD4値（確認用）
        # print(PD2kosuExcel_PersonalSheet["D4"].value)
        # 出勤時間を入れる
        PD2kosuExcel_PersonalSheet["C{}".format(int(date02) + 8)].value = None
        # 退勤時間を入れる

        PD2kosuExcel_PersonalSheet["E{}".format(int(date02) + 8)].value = None
        # 休憩開始時間を入れる

        PD2kosuExcel_PersonalSheet["F{}".format(int(date02) + 8)].value = None
            # 休憩終了時間を入れる
        PD2kosuExcel_PersonalSheet["G{}".format(int(date02) + 8)].value = None
        # 　食事数を入れる。
        break

JissekiExcelSheet = file2[NPdate]
for cells in JissekiExcelSheet["D"]:
    print(cells)
    print(cells.value)
    print(str(NPnumber))
    if str(cells.value) == str(NPnumber):
        print(cells.value)
        JissekiExcelSheet.cell(row = cells.row, column = 8).value = "削除済み"
        JissekiExcelSheet.cell(row = cells.row, column = 9).value = "削除済み"
        JissekiExcelSheet.cell(row = cells.row, column = 10).value = "削除済み"
        break

file.save(path1)
file2.save(path2)