import openpyxl
path1 = r"C:\1\PD2\PD2工数.xlsx"
file = openpyxl.load_workbook(path1)
sheet0 = file["0SHEET"]
sheet2 = file["2在籍者名簿Master"]

NPnumber = input("4桁番号を入力してください")
NPname = input("名前を入力してください")
for cell in sheet0["D"]:
    if cell.value == None:
        print(cell.row)
        sheet0.cell(row=cell.row,column=3).value = int(NPnumber)
        cell.value = NPname
        break
for cell in sheet2["A"]:
    if cell.value == None:
        cell.value = int(NPnumber)
        sheet2.cell(row=cell.row,column=2).value = NPname
        break


newsheet = file.copy_worksheet(file.worksheets[3])
newsheet.title=str(NPnumber) + " " + NPname
newsheet["D4"] = NPname
newsheet["I4"] = NPnumber
newsheet.sheet_view.showZeros = False
file.move_sheet(newsheet,-1)
file.save(path1)