import openpyxl
from tkinter import filedialog
from datetime import timedelta

file_path = filedialog.askopenfilename()
file = openpyxl.load_workbook(file_path,data_only=True)
workbook = file.active
sheetlen = len(workbook["A"])
listSN = []
listKN = []
NumCountSN = 0
NumCountKN = 0
WorktimeCountSN = 0
WorktimeCountKN = 0
wordSN = "湘南"
wordKN = "神奈川"

for row in range(2,sheetlen+1):
    Factory_Name = workbook.cell(row=row,column=5).value
    if Factory_Name == None:
        break
    print(Factory_Name)
    if wordSN in Factory_Name:
        NumCountSN += 1
        for col in range(11,13):
            time_nowRow = workbook.cell(row=row,column=col).value
            WorktimeCountSN = WorktimeCountSN + time_nowRow
    elif wordKN in Factory_Name:
        NumCountKN += 1
        for col in range(11,13):
            time_nowRow = workbook.cell(row=row,column=col).value
            WorktimeCountKN = WorktimeCountKN + time_nowRow
    else:
        pass

print("工数集計　湘南工場　人数：{}　時間数：{}　　　　\n神奈川工場　人数：{}　時間数:{}".format(NumCountSN,WorktimeCountSN,NumCountKN,WorktimeCountKN))