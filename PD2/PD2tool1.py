import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.views
from openpyxl.worksheet.views import SheetView

filepath = (r"C:\Users\timaz\Documents\PythonFile\PD2\PD2工数.xlsx")
file = openpyxl.load_workbook(filepath)
print(file)
# for x in file.worksheets:
#     print(x)

AllNameListSheet = file["2在籍者名簿Master"]
namelist = []
foodsheet = file["0SHEET"]
print_area = 'A1:Q41'


for i in range (1,AllNameListSheet.max_row+1):
    valueNow = AllNameListSheet.cell(row=i,column=1).value
    nameNow = AllNameListSheet.cell(row=i,column=2).value
    x = [valueNow,nameNow]
    if valueNow == None or valueNow == "":
        break
    else:
        newsheet = file.copy_worksheet(file.worksheets[3])
        newsheet.title=str(valueNow-70000000) + " " + nameNow
        newsheet["D4"] = nameNow
        newsheet["I4"] = int(valueNow) - 70000000
        newsheet.print_area = print_area
        # 横を１ページ
        newsheet.fitToWidth = 1
        # 縦を自動
        newsheet.fitToHeight = 1
        newsheet.sheet_properties.pageSetUpPr.fitToPage = True
        newsheet.sheet_view.showZeros = False
              
        
        file.move_sheet(newsheet,-1)
        foodsheet.cell(row = i+3,column=3,value= valueNow-70000000)
        foodsheet.cell(row = i+3,column=4,value = nameNow)







for xx in namelist:
    print(xx)

    
    # print(valueNow)
    # if valueNow == None or valueNow == "":
    #     print("skip")

print(AllNameListSheet)

file.save("test.xlsx")
file.close