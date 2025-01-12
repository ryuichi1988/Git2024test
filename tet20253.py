import cv2
import pytesseract
import pandas as pd

# 设置 Tesseract 路径
pytesseract.pytesseract.tesseract_cmd = r'C:\Users\timaz\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

# 加载图像
image_path = r'C:\1\2.jpg'
image = cv2.imread(image_path)

# 初始化参数
drawing = False  # 是否正在绘制
x_start, y_start = -1, -1  # 起始点
regions = {}  # 存储选择的区域

# 鼠标回调函数
def draw_rectangle(event, x, y, flags, param):
    global x_start, y_start, drawing, img_copy, regions
    if event == cv2.EVENT_LBUTTONDOWN:  # 鼠标左键按下
        drawing = True
        x_start, y_start = x, y
    elif event == cv2.EVENT_MOUSEMOVE:  # 鼠标移动
        if drawing:
            img_copy = image.copy()
            cv2.rectangle(img_copy, (x_start, y_start), (x, y), (0, 255, 0), 2)
            cv2.imshow("Select Region", img_copy)
    elif event == cv2.EVENT_LBUTTONUP:  # 鼠标左键释放
        drawing = False
        region_name = f"field_{len(regions)+1}"  # 自动生成区域名称
        regions[region_name] = (x_start, y_start, x, y)  # 保存矩形区域坐标
        print(f"Region {region_name}: ({x_start}, {y_start}, {x}, {y})")  # 输出矩形区域坐标

# 创建窗口并绑定回调函数
cv2.namedWindow("Select Region")
cv2.setMouseCallback("Select Region", draw_rectangle)

# 显示图像并等待用户手动选择区域
img_copy = image.copy()
while True:
    cv2.imshow("Select Region", img_copy)
    key = cv2.waitKey(1) & 0xFF
    if key == 27:  # 按 ESC 键退出
        break

# 关闭窗口
cv2.destroyAllWindows()

# 提取 OCR 数据
extracted_data = {}
for field_name, (x1, y1, x2, y2) in regions.items():
    # 裁剪图像，只保留选定的区域
    region = image[y1:y2, x1:x2]
    
    # 显示裁剪的区域（调试用）
    cv2.imshow(f"{field_name}", region)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

    # OCR 提取文字，使用日语语言模型
    config = '--psm 6'  # 单行模式
    text = pytesseract.image_to_string(region, lang='jpn', config=config)
    
    # 提取表格数据（此时可以进一步处理格式化）
    rows = text.split("\n")
    table_data = [row.split() for row in rows if row.strip()]  # 按行分割，去除空行
    
    extracted_data[field_name] = table_data

# 打印提取结果
for field, value in extracted_data.items():
    print(f"{field}:")
    for row in value:
        print(row)

# 将结果写入 Excel 文件
# 使用 openpyxl 库将结果写入 Excel
from openpyxl import Workbook

# 创建一个新的工作簿
wb = Workbook()
# 遍历每个区域，创建对应的工作表
for field_name, table in extracted_data.items():
    ws = wb.create_sheet(title=field_name)  # 创建新的工作表
    for row_index, row in enumerate(table, start=1):
        for col_index, cell in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=cell)  # 填充单元格

# 删除默认生成的空工作表
del wb['Sheet']

# 保存到 Excel 文件
wb.save('extracted_table.xlsx')
print("Table data has been written to 'extracted_table.xlsx'.")
