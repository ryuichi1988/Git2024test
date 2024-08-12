"""
# 「名前D、　番号B、　出勤時間n、　作業ラインn+1、　部署H。」


2021.7.26更新：
詳細取得時、範囲を修正しました
（第１２７行コード）↓
×　2:147は間違い。
☑　2:149 (149は範囲として取らない。)

2021.8.23更新：
新人が増えたので、範囲を修正しました。
（第１２７行コード）↓
for row in range(2, 152):


2021.9.11更新：
性別欄自動追加できるようになりました。



2022.4.30　4.5UPDATE
ルメールさん　食事×

2022.7.29　4.6UPDATE
ルメールさん　食事○



2024.3.5
①もし番号のところは数字以外⇒数字の判定式修正。
②固定ライン表示仕方、順番並びルール変更。
取り方⇒盛付の順
"""

# import PySimpleGUI as sg　　※　v4.1以降：PySimpleGUIはたまにフリーズするので、使用中止。twinkerに変更しました。


import datetime
import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from os import system
from openpyxl.styles.fonts import Font

root = tk.Tk()
root.withdraw()
font = Font(size = 24, bold = True)

# 获取文件路径
file_path = filedialog.askopenfilename()
print(file_path)


# list第２要素並べ替え
def takeSecond(elem):
    return elem[2]


# list第５要素並べ替え
def takeFourth(elem):
    return elem[5]


# sg.theme('Default')

# 時間リスト作成
list83 = []
list10 = []
list15 = []
list17 = []
list11 = []

# 現在と明日
now = datetime.datetime.now()
tomorrow = now + datetime.timedelta(days=1)

print("tomorrow type:")
print(type(tomorrow))  # type(tomorrow) == datetime.datetime

filename = ("ニッセープロダクツ出勤簿{}月{}日.xlsx".format(tomorrow.month, tomorrow.day))

# show1 = "{:02}{:02}".format(tomorrow.month, tomorrow.day)
# sg.theme('Dark Blue 3')   # Add a touch of color
# All the stuff inside your window.
# layout = [[sg.Text('関東ダイエットクック神奈川工場')],
#           [sg.Text('日付を選択してください。'), sg.InputText(default_text=show1, size=(10, 1))],
#           [sg.Text("ファイル"), sg.InputText(size=(85, 1)), sg.FileBrowse(key="file1", button_text="ファイルを選択")],
#           [sg.Submit(button_text="テストボタン", key="Submit")],
#           [sg.Button('出勤簿を作成'), sg.Cancel(button_text="EXIT")]]

# Create the Window
# window = sg.Window('NisseyProducts出勤簿作成ツール　DC神奈川工場用3.1', layout)
# Event Loop to process "events" and get the "values" of the inputs


test = datetime.time()
outputfile = openpyxl.load_workbook(r"C:\Users\timaz\Documents\PythonFile\関東ダイエットクック手書き出勤簿.xlsx")
OutPut__sheets = outputfile.worksheets[0]

# 日付を直す。L1　＝　明日
OutPut__sheets["M1"] = tomorrow

# 空リスト
emptyLIST = []
# while True:
#     event, values = window.read()
#     if event == "出勤簿を作成":
#         print("作成中....")
#         if len(values["file1"]) < 5:
#             print("file:{}".format(values["file1"]))
#             print("FIlEが選択されていません。")
#             continue


# 　★

main_excel_file = openpyxl.load_workbook(file_path, data_only=True)
sheet = main_excel_file.active
for row in range(1, 2):  # 一列目指定
    for col in range(1, 111):  # 日付遍歴
        # print(sheet.cell(column=col,row=row).value)
        cellnow = sheet.cell(column=col, row=row).value  # 日付セル値プリント
        cellnoww = sheet.cell(column=col, row=row)  # 日付セル特定プリント
        print("-----------------------")
        print("第{}列".format(col))
        print("checking:")
        print(cellnoww)
        print(cellnow)
        print("with...")
        print(tomorrow)
        print("cellnow w types:")
        print(type(cellnow))
        print(type(cellnoww))

        if type(cellnow) == datetime.datetime:  # datetime.datetimeか
            print("これはデータタイムです。")
        else:
            print("これはデータタイムではありません。")
            continue
        try:
            if cellnow.month == tomorrow.month and cellnow.day == tomorrow.day:  # もし日付ヒットしましたら。
                print("★★★★yes:↓★★★★★★★★★★★★★★★★★★★★★★★★")
                print(cellnow)

                # ★★★★★★★左取右不取哦  这个是从第几排到第几排★★★★★★★
                for row in range(2, 170):
                    # ★★★★★★★左取右不取哦  这个是从第几排到第几排★★★★★★★

                    names = sheet.cell(row=row, column=4).value
                    number = sheet.cell(row=row, column=1).value
                    sex = sheet.cell(row=row, column=5).value

                    if number:
                        pass
                    else:
                        number = 9999


                    if str(number) == "":
                        number = 9999
                    else:
                        pass  # もし番号が数字でないなら、既定は9999。


                    start_time = sheet.cell(row=row, column=col).value
                    work_line = sheet.cell(row=row, column=col + 1).value
                    department = sheet.cell(row=row, column=8).value  # 第８列：H
                    if department is None:
                        department = "盛付"  # もし部署が入力されていないなら、既定は盛付。
                    else:
                        pass

                    names2 = sheet.cell(row=row, column=4)
                    sex2 = sheet.cell(row=row, column=5)
                    number2 = sheet.cell(row=row, column=1)
                    start_time2 = sheet.cell(row=row, column=col)
                    work_line2 = sheet.cell(row=row, column=col + 1)
                    department2 = sheet.cell(row=row, column=8)  # 第８列：H列
                    #固定メンバーのライン表示
                    if department == "盛付A":
                        work_line = "A"
                        department = "盛付"
                    
                    if department == "盛付B":
                        work_line = "B"
                        department = "盛付"

                    if department == "盛付C":
                        work_line = "C"
                        department = "盛付"

                    if department == "盛付D":
                        work_line = "D"
                        department = "盛付"

                    if department == "盛付E":
                        work_line = "E"
                        department = "盛付"

                    if department == "盛付F":
                        work_line = "F"
                        department = "盛付"

                    if department == "盛付G":
                        work_line = "G"
                        department = "盛付"

                    if department == "盛付I":
                        work_line = "I"
                        department = "盛付"

                    if department == "盛付J":
                        work_line = "J"
                        department = "盛付"

                   
                    # 取り方及び、オペレーターの処理。
                    if work_line == "T":
                        work_line = ""
                        department = "取り方"
                    # elif work_line == "焼き物室":
                    #     work_line = "YAKI"
                    #     department = "焼き物"
                    # elif work_line == "YAKI":
                    #     work_line = ""
                    #     department = "焼き物"
                    elif work_line == "o" or work_line == "O":
                        work_line = ""
                        department = "オペレーター"
                    else:
                        pass

                    if type(start_time) != datetime.time:
                        continue
                    print("{},{},{},{},{},{}".format(names2, sex2, number2, start_time2, work_line2, department2))
                    print("{},{},{},{},{},{}".format(names, sex, number, start_time, work_line, department))
                    # 「名前D、　番号B、　出勤時間n、　作業ラインn+1、　部署H。」
                    print("---------------------------")

                    Row__Now = [names, sex, int(number), start_time, work_line, department]
                    emptyLIST.append(Row__Now)

                break  # もしヒットしたらもちろんBREAK。
            else:
                print("check finish, but no match for date.ERROR")
                print("☑終了だが、日付一致のセルがありませんでした。")

        except TypeError or AttributeError:
            print("TypeError,AttribERROR")
            print(cellnow)
            print(cellnow.value)
            print(cellnoww)
            print(cellnoww.value)
            continue
        print("-----------------------\n")
    print("リスト第[1]を並べ替え")
    for iii in emptyLIST:
        print(iii[2])
    emptyLIST.sort(key=takeSecond)  # リスト第[１]を並べ替え
    print("リスト第[4]を並べ替え")
    for iii in emptyLIST:
        print(iii[5])
    emptyLIST.sort(key=takeFourth)  # リスト第[４]を並べ替え [names, sex, int(number), start_time, work_line, department]
    for i in emptyLIST:
        print(i)
        if i[3] == datetime.time(8, 30):
            list83.append(i)
        elif i[3] == datetime.time(10, 00):
            list10.append(i)
        elif i[3] == datetime.time(11, 00):
            list11.append(i)
        elif i[3] == datetime.time(15, 00):
            list15.append(i)
        elif i[3] == datetime.time(17, 00):
            list17.append(i)
        else:
            print(i)
            print("???")
            list83.append(i)

    print(len(emptyLIST))
    print(len(list83))
    print(len(list10))
    print(len(list11))
    print(len(list15))
    print(len(list17))
    print("000000000\n")

    for iii in emptyLIST:
        print(iii)

#   for row in range(1,len(list) + 10):
# EXCELに入力し始める。↓まずＲｏｗカーソルを１０にセットします。
row = 10
for i in list83:
    OutPut__sheets.cell(row=row, column=2).value = i[2]  # [names, sex, int(number), start_time, work_line, department]
    OutPut__sheets.cell(row=row, column=4).value = i[0]  # name
    OutPut__sheets.cell(row=row, column=3).value = i[1]  # sex
    if i[4] == "YAKI":
        OutPut__sheets.cell(row=row, column=6).value = "YAKI"  # dep
        OutPut__sheets.cell(row=row, column=5).value = ""  # li-ne
    elif i[4] == "キット室":
        OutPut__sheets.cell(row=row, column=6).value = "キット室"  # dep
        OutPut__sheets.cell(row=row, column=5).value = ""  # li-ne
    else:
        OutPut__sheets.cell(row=row, column=6).value = i[5]
        OutPut__sheets.cell(row=row, column=5).value = i[4]  # li-ne




    OutPut__sheets.cell(row=row, column=7).value = i[3]  # sta
    OutPut__sheets.row_dimensions[row].height = 28
    row += 1

row += 2
for i in list11:
    OutPut__sheets.cell(row=row, column=2).value = i[2]  # [names, sex, int(number), start_time, work_line, department]
    OutPut__sheets.cell(row=row, column=4).value = i[0]  # name
    OutPut__sheets.cell(row=row, column=3).value = i[1]  # sex
    OutPut__sheets.cell(row=row, column=6).value = i[5]  # dep
    OutPut__sheets.cell(row=row, column=5).value = i[4]  # li-ne
    OutPut__sheets.cell(row=row, column=7).value = i[3]  # sta
    OutPut__sheets.cell(row=row, column=11).value = "○"
    if i[2] == 6663:
        OutPut__sheets.cell(row=row, column=11).value = "×"
    else:
        pass
    OutPut__sheets.cell(row=row, column=11).font = font
    row += 1

row += 1
for _ in range(17):
    if str(OutPut__sheets.cell(row=row, column=1).value).isdigit():
        OutPut__sheets.row_dimensions[row].hidden = True
    row += 1

# OutPut__sheets.column_dimensions["C"].width = 28
OutPut__sheets = outputfile.worksheets[1]
row = 10
for i in list10:
    OutPut__sheets.cell(row=row, column=2).value = i[2]  # [names, sex, int(number), start_time, work_line, department]
    OutPut__sheets.cell(row=row, column=4).value = i[0]  # name
    OutPut__sheets.cell(row=row, column=3).value = i[1]  # sex
    OutPut__sheets.cell(row=row, column=6).value = i[5]  # dep
    OutPut__sheets.cell(row=row, column=5).value = i[4]  # li-ne
    OutPut__sheets.cell(row=row, column=7).value = i[3]  # sta
    OutPut__sheets.cell(row=row, column=11).value = "○"
    row += 1
# OutPut__sheets.column_dimensions["C"].width = 28

print('''
OutPut__sheets = outputfile.worksheets[2]
row = 10
for i in list15:
    OutPut__sheets.cell(row=row, column=2).value = i[2]  # [names, sex, int(number), start_time, work_line, department]
    OutPut__sheets.cell(row=row, column=4).value = i[0]  # name
    OutPut__sheets.cell(row=row, column=3).value = i[1]  # sex
    OutPut__sheets.cell(row=row, column=6).value = i[5]  # dep
    OutPut__sheets.cell(row=row, column=5).value = i[4]  # li-ne
    OutPut__sheets.cell(row=row, column=7).value = i[3]  # sta
    OutPut__sheets.cell(row=row, column=11).value = "×"
    OutPut__sheets.cell(row=row, column=10).value = "×"

    row += 1
for i in list17:
    OutPut__sheets.cell(row=row, column=2).value = i[2]  # [names, sex, int(number), start_time, work_line, department]
    OutPut__sheets.cell(row=row, column=4).value = i[0]  # name
    OutPut__sheets.cell(row=row, column=3).value = i[1]  # sex
    OutPut__sheets.cell(row=row, column=6).value = i[5]  # dep
    OutPut__sheets.cell(row=row, column=5).value = i[4]  # li-ne
    OutPut__sheets.cell(row=row, column=7).value = i[3]  # sta
    OutPut__sheets.cell(row=row, column=11).value = "×"
    OutPut__sheets.cell(row=row, column=10).value = "×"

    row += 1

# OutPut__sheets.column_dimensions["C"].width = 28

''')

outputfile.save(filename)

main_excel_file.close()
outputfile.close()
print("file saved.")


# messagebox.showinfo("㈱ニッセープロダクツ　㈱関東ダイエットクック出勤簿V4.0 by 奥山","{}月{}日出勤簿が作成しました。\n\n{}\n\nOKボタンを押すとファイルが開く。"\
#                     .format(tomorrow.month, tomorrow.day, filename))


system(filename)
#
#
#
# if event == "Submit":
#     now2 = datetime.datetime.now()
#     print('{}あなたばテストボタンを押しましたね。何をするつもりですか。'.format(now2), values[0])
#
# if event == sg.WIN_CLOSED or event == 'EXIT':  # if user closes window or clicks cancel
#     try:

# except NameError:
#     print("NameError Exiting 1..")
#     break
# print("Normal Exiting 0..")
# break

# print(values["file1"])

# window.close()
